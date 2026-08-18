#!/usr/bin/env python3
"""Focused backend/API verification for RTO Total Deductions and DTO/RTO signs.

Runs against Preview only. It verifies the exact Google Doc bug:
RTO Total Deductions must equal Commission + Fixed Fee + GT + Return Fee,
while preserving DTO/RTO sign conventions and expected_settlement semantics.
"""

import io
import json
import math
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests


BASE_URL = os.environ.get("PREVIEW_BASE_URL", "https://settlement-intel-1.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"
OUT = Path("/app/test_reports/bug_verify_rto_totals_iter24_backend_result.json")


def near(a, b, tol=0.02):
    return math.isclose(float(a), float(b), abs_tol=tol)


def money(v):
    if v is None or v == "":
        return None
    return round(float(v), 2)


class CheckFailure(AssertionError):
    pass


class Runner:
    def __init__(self):
        self.s = requests.Session()
        self.evidence = {}
        self.failures = []

    def request(self, method, path, **kwargs):
        url = f"{API}{path}"
        r = self.s.request(method, url, timeout=60, **kwargs)
        if not r.ok:
            raise CheckFailure(f"{method} {path} failed {r.status_code}: {r.text[:500]}")
        return r

    def login(self):
        r = self.request("POST", "/auth/login", json={"email": "admin@fundle.ai", "password": "admin123"})
        token = r.json().get("token")
        if token:
            self.s.headers.update({"Authorization": f"Bearer {token}"})
        self.evidence["auth"] = {"logged_in_as": r.json().get("user", {}).get("email"), "has_token": bool(token)}

    def get_calculations(self, order_type, limit=200):
        r = self.request(
            "GET",
            "/calculations",
            params={
                "portal": "myntra",
                "order_type": order_type,
                "severity_flag": "mapped",
                "limit": limit,
                "sort_by": "computed_at",
                "sort_dir": "desc",
            },
        )
        data = r.json()
        rows = data.get("items") or []
        if data.get("total", 0) < 5 or len(rows) < 5:
            raise CheckFailure(f"Expected at least 5 {order_type} rows, got total={data.get('total')} items={len(rows)}")
        self.evidence[f"{order_type}_total"] = data.get("total")
        return rows

    def sample_rows(self, rows, order_type, predicate, needed=5):
        samples = [r for r in rows if predicate(r)]
        if len(samples) < needed:
            preview = [
                {
                    "order_id": r.get("online_order_id"),
                    "commission": r.get("commission_incl_gst"),
                    "fixed": r.get("fixed_fee_incl_gst"),
                    "gt": r.get("gt_charge"),
                    "return_fee": r.get("return_fee"),
                    "total_deductions": r.get("total_deductions"),
                    "expected_settlement": r.get("expected_settlement"),
                }
                for r in rows[:10]
            ]
            raise CheckFailure(f"Could not find {needed} valid {order_type} sample rows. Preview={preview}")
        return samples[:needed]

    def component_sum(self, r):
        return round(
            money(r.get("commission_incl_gst"))
            + money(r.get("fixed_fee_incl_gst"))
            + money(r.get("gt_charge"))
            + money(r.get("return_fee")),
            2,
        )

    def check_rto(self):
        rows = self.get_calculations("rto")
        samples = self.sample_rows(
            rows,
            "rto",
            lambda r: all(money(r.get(k)) is not None for k in ["commission_incl_gst", "fixed_fee_incl_gst", "gt_charge", "return_fee", "total_deductions", "expected_settlement"]),
        )
        checked = []
        for r in samples:
            comm = money(r["commission_incl_gst"])
            fixed = money(r["fixed_fee_incl_gst"])
            gt = money(r["gt_charge"])
            ret = money(r["return_fee"])
            td = money(r["total_deductions"])
            es = money(r["expected_settlement"])
            summed = self.component_sum(r)
            if not (comm < 0 and fixed < 0 and gt < 0 and ret < 0):
                raise CheckFailure(f"RTO signs wrong for {r.get('online_order_id')}: comm={comm}, fixed={fixed}, gt={gt}, return={ret}")
            if not near(td, summed):
                raise CheckFailure(f"RTO total_deductions mismatch for {r.get('online_order_id')}: td={td}, sum={summed}")
            if not near(es, 0):
                raise CheckFailure(f"RTO expected_settlement should be 0 for {r.get('online_order_id')}: {es}")
            checked.append({"order_id": r.get("online_order_id"), "month": r.get("report_month"), "commission": comm, "fixed_fee": fixed, "gt": gt, "return_fee": ret, "sum": summed, "total_deductions": td, "expected_settlement": es})
        self.evidence["rto_samples"] = checked
        return samples

    def check_return_dto(self):
        rows = self.get_calculations("return_dto")
        samples = self.sample_rows(
            rows,
            "return_dto",
            lambda r: all(money(r.get(k)) is not None for k in ["commission_incl_gst", "fixed_fee_incl_gst", "gt_charge", "return_fee", "total_deductions"]),
        )
        checked = []
        for r in samples:
            comm = money(r["commission_incl_gst"])
            fixed = money(r["fixed_fee_incl_gst"])
            gt = money(r["gt_charge"])
            ret = money(r["return_fee"])
            td = money(r["total_deductions"])
            summed = self.component_sum(r)
            if not (comm < 0 and fixed == 0 and gt < 0 and ret > 0):
                raise CheckFailure(f"return_dto signs wrong for {r.get('online_order_id')}: comm={comm}, fixed={fixed}, gt={gt}, return={ret}")
            if not near(td, summed):
                raise CheckFailure(f"return_dto total_deductions mismatch for {r.get('online_order_id')}: td={td}, sum={summed}")
            checked.append({"order_id": r.get("online_order_id"), "commission": comm, "fixed_fee": fixed, "gt": gt, "return_fee": ret, "sum": summed, "total_deductions": td})
        self.evidence["return_dto_samples"] = checked

    def check_sales_and_return_regression(self):
        regression = {}
        for order_type in ["sales", "return"]:
            rows = self.get_calculations(order_type)
            samples = self.sample_rows(
                rows,
                order_type,
                lambda r: all(money(r.get(k)) is not None for k in ["commission_incl_gst", "fixed_fee_incl_gst", "gt_charge", "return_fee", "total_deductions", "expected_settlement", "nsv_after_gt"]),
            )
            checked = []
            for r in samples:
                comm = money(r["commission_incl_gst"])
                fixed = money(r["fixed_fee_incl_gst"])
                gt = money(r["gt_charge"])
                ret = money(r["return_fee"])
                td = money(r["total_deductions"])
                nsv_after_gt = money(r["nsv_after_gt"])
                es = money(r["expected_settlement"])
                summed = self.component_sum(r)
                if not near(td, summed):
                    raise CheckFailure(f"{order_type} total_deductions mismatch for {r.get('online_order_id')}: td={td}, sum={summed}")
                if not near(es, round(nsv_after_gt - td, 2)):
                    raise CheckFailure(f"{order_type} expected_settlement mismatch for {r.get('online_order_id')}: es={es}, nsv_after_gt-td={round(nsv_after_gt - td, 2)}")
                if order_type == "sales" and not (comm >= 0 and fixed >= 0 and gt >= 0 and ret == 0):
                    raise CheckFailure(f"sales signs regressed for {r.get('online_order_id')}: comm={comm}, fixed={fixed}, gt={gt}, return={ret}")
                if order_type == "return" and not (comm < 0 and fixed < 0 and gt < 0 and ret > 0):
                    raise CheckFailure(f"return signs regressed for {r.get('online_order_id')}: comm={comm}, fixed={fixed}, gt={gt}, return={ret}")
                checked.append({"order_id": r.get("online_order_id"), "commission": comm, "fixed_fee": fixed, "gt": gt, "return_fee": ret, "sum": summed, "total_deductions": td, "expected_settlement": es})
            regression[order_type] = checked
        self.evidence["sales_return_regression_samples"] = regression

    def check_export(self, rto_samples):
        month = next((r.get("report_month") for r in rto_samples if r.get("report_month")), None)
        if not month:
            raise CheckFailure("Could not determine report_month from RTO samples for export test")
        r = self.request("GET", "/sales/export", params={"portal": "myntra", "report_month": month, "order_status": "RTO"})
        content_type = r.headers.get("content-type", "")
        if "spreadsheet" not in content_type and not r.content.startswith(b"PK"):
            raise CheckFailure(f"Export did not return an XLSX. content-type={content_type}, prefix={r.content[:50]!r}")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}
        needed = ["Order Status", "Online Order ID", "Commission (Expected)", "GT Charge (Expected)", "Fixed Fee (Expected)", "Return Fee (Expected)", "Total Deductions", "Expected Settlement"]
        missing = [h for h in needed if h not in idx]
        if missing:
            raise CheckFailure(f"Export missing headers: {missing}; headers={headers}")
        checked = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[idx["Order Status"]] or "").strip().upper() != "RTO":
                continue
            comm = money(row[idx["Commission (Expected)"]])
            fixed = money(row[idx["Fixed Fee (Expected)"]])
            gt = money(row[idx["GT Charge (Expected)"]])
            ret = money(row[idx["Return Fee (Expected)"]])
            td = money(row[idx["Total Deductions"]])
            es = money(row[idx["Expected Settlement"]])
            if None in (comm, fixed, gt, ret, td):
                continue
            summed = round(comm + fixed + gt + ret, 2)
            if not near(td, summed):
                raise CheckFailure(f"Export RTO total mismatch for {row[idx['Online Order ID']]}: td={td}, sum={summed}")
            if not (comm < 0 and fixed < 0 and gt < 0 and ret < 0):
                raise CheckFailure(f"Export RTO signs wrong for {row[idx['Online Order ID']]}: comm={comm}, fixed={fixed}, gt={gt}, return={ret}")
            if not near(es, 0):
                raise CheckFailure(f"Export RTO expected_settlement should be 0 for {row[idx['Online Order ID']]}: {es}")
            checked.append({"order_id": row[idx["Online Order ID"]], "commission": comm, "fixed_fee": fixed, "gt": gt, "return_fee": ret, "sum": summed, "total_deductions": td, "expected_settlement": es})
            if len(checked) >= 5:
                break
        if len(checked) < 5:
            raise CheckFailure(f"Export had fewer than 5 RTO rows for month={month}; checked={len(checked)}")
        self.evidence["export"] = {"month": month, "checked_rto_rows": checked, "content_type": content_type, "bytes": len(r.content)}

    def run(self):
        status = "passed"
        try:
            self.login()
            rto_samples = self.check_rto()
            self.check_return_dto()
            self.check_export(rto_samples)
            self.check_sales_and_return_regression()
        except Exception as e:
            status = "failed"
            self.failures.append(str(e))
        result = {
            "status": status,
            "base_url": BASE_URL,
            "ran_at": datetime.now(timezone.utc).isoformat(),
            "evidence": self.evidence,
            "failures": self.failures,
        }
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps(result, indent=2), encoding="utf-8")
        print(json.dumps(result, indent=2))
        return 0 if status == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(Runner().run())