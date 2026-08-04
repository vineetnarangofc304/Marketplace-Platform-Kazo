#!/usr/bin/env python3
"""Focused backend/API verification for the Sales Ledger client observations.

This script is intentionally narrow: it logs in, verifies the April 2026 Myntra
Sales Ledger summary/export, verifies DTO/sales calculation arithmetic, runs the
recalculation endpoint requested by the fix, and checks the commission ALL
fallback evidence through the public API.
"""
import io
import json
import math
import os
import re
import sys
from pathlib import Path

import openpyxl
import requests


ROOT = Path("/app")
OUT = ROOT / "test_reports" / "test_sales_ledger_backend_results.json"


def read_backend_url() -> str:
    env_text = (ROOT / "frontend" / ".env").read_text()
    m = re.search(r"^REACT_APP_BACKEND_URL=(.+)$", env_text, re.M)
    if not m:
        raise RuntimeError("REACT_APP_BACKEND_URL not found")
    return m.group(1).strip().strip('"')


def almost(a, b, tol=0.02):
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tol


def nz(v):
    return 0 if v is None else v


def api_get(sess, api_base, path, **kwargs):
    r = sess.get(api_base + path, timeout=90, **kwargs)
    r.raise_for_status()
    return r


def main():
    base = read_backend_url()
    api = base.rstrip("/") + "/api"
    sess = requests.Session()
    results = {
        "api_base": api,
        "checks": [],
        "failures": [],
        "samples": {},
    }

    def check(name, ok, detail):
        results["checks"].append({"name": name, "ok": bool(ok), "detail": detail})
        print(("PASS" if ok else "FAIL") + f" - {name}: {detail}")
        if not ok:
            results["failures"].append({"name": name, "detail": detail})

    # Auth
    login_payloads = [
        {"email": "admin@fundle.ai", "password": "admin123"},
        {"email": "admin@kazo.com", "password": "admin123"},
    ]
    login_data = None
    for payload in login_payloads:
        r = sess.post(api + "/auth/login", json=payload, timeout=30)
        if r.status_code == 200:
            login_data = r.json()
            sess.headers.update({"Authorization": f"Bearer {login_data['token']}"})
            break
    check("admin login", login_data is not None, login_data.get("user") if login_data else "login failed")
    if login_data is None:
        raise SystemExit(2)

    params_apr = {"period_type": "month", "period_value": "2026-04", "portal": "myntra"}

    # Sales Ledger summary: net orders = Sales - Returns, expected April values.
    summary = api_get(sess, api, "/sales/summary", params=params_apr).json()
    results["samples"]["sales_summary"] = summary
    expected_net = summary.get("sales_rows", 0) - summary.get("return_rows", 0)
    check(
        "sales summary net_orders equals sales_rows - return_rows",
        summary.get("net_orders") == expected_net,
        summary,
    )
    check(
        "sales summary April 2026 expected values",
        summary.get("sales_rows") == 14219 and summary.get("return_rows") == 7395 and summary.get("net_orders") == 6824,
        summary,
    )

    # Sales Ledger rows expose requested source columns.
    sales_rows = api_get(sess, api, "/sales", params={**params_apr, "limit": 20}).json()
    rows = sales_rows.get("items", [])
    source_fields = ["brand", "txn_type", "posting_date", "sku", "posting_location_code", "main_category", "sub_category", "zone", "report_month", "qty", "mrp", "nsv_val"]
    populated_counts = {f: sum(1 for row in rows if row.get(f) not in (None, "")) for f in source_fields}
    check(
        "sales API first 20 include requested source fields",
        len(rows) >= 20 and all(populated_counts[f] > 0 for f in source_fields),
        {"row_count": len(rows), "populated_counts": populated_counts},
    )

    # DTO returns: negative reversals, positive return fee, no GST/TCS/TDS.
    dto_rows = api_get(sess, api, "/calculations", params={**params_apr, "order_type": "return_dto", "limit": 50}).json().get("items", [])
    dto_failures = []
    for c in dto_rows[:20]:
        if not (c.get("commission_incl_gst") is not None and c.get("commission_incl_gst") < 0):
            dto_failures.append((c.get("id"), "commission_incl_gst not negative", c.get("commission_incl_gst")))
        if not (c.get("gt_charge") is not None and c.get("gt_charge") < 0):
            dto_failures.append((c.get("id"), "gt_charge not negative", c.get("gt_charge")))
        if not (c.get("fixed_fee_incl_gst") is not None and c.get("fixed_fee_incl_gst") < 0):
            dto_failures.append((c.get("id"), "fixed_fee_incl_gst not negative", c.get("fixed_fee_incl_gst")))
        if not (c.get("return_fee") is not None and c.get("return_fee") > 0):
            dto_failures.append((c.get("id"), "return_fee not positive", c.get("return_fee")))
        for z in ["tcs", "tds", "commission_gst", "fixed_fee_gst"]:
            if c.get(z) != 0:
                dto_failures.append((c.get("id"), f"{z} not zero", c.get(z)))
    if dto_rows:
        results["samples"]["return_dto"] = dto_rows[0]
    check("return_dto calculations have reversals and no tax/GST", bool(dto_rows) and not dto_failures, {"sample_count": len(dto_rows), "failures": dto_failures[:5]})

    # Sales rows: TCS/TDS/GST zero and settlement formula.
    sales_calcs = api_get(sess, api, "/calculations", params={**params_apr, "order_type": "sales", "limit": 50}).json().get("items", [])
    sales_failures = []
    for c in sales_calcs[:20]:
        for z in ["tcs", "tds", "commission_gst", "fixed_fee_gst"]:
            if c.get(z) != 0:
                sales_failures.append((c.get("id"), f"{z} not zero", c.get(z)))
        if not almost(c.get("commission_incl_gst"), c.get("commission_base")):
            sales_failures.append((c.get("id"), "commission_incl_gst != commission_base", c.get("commission_incl_gst"), c.get("commission_base")))
        if c.get("return_fee") not in (0, 0.0):
            sales_failures.append((c.get("id"), "sales return_fee not zero", c.get("return_fee")))
        expected_settlement = nz(c.get("nsv_after_gt")) - (nz(c.get("commission_incl_gst")) + nz(c.get("fixed_fee_incl_gst")) + nz(c.get("gt_charge")))
        if not almost(c.get("expected_settlement"), expected_settlement):
            sales_failures.append((c.get("id"), "settlement formula mismatch", c.get("expected_settlement"), expected_settlement))
    if sales_calcs:
        results["samples"]["sales_calc"] = sales_calcs[0]
    check("sales calculations remove TCS/TDS/GST and match formula", bool(sales_calcs) and not sales_failures, {"sample_count": len(sales_calcs), "failures": sales_failures[:5]})

    # Run recalculation for the month/portal and verify unmapped disappears.
    run_resp = sess.post(api + "/calculations/run", json={"report_month": "2026-04", "portal": "myntra", "recalculate": True}, timeout=240)
    run_json = run_resp.json() if run_resp.content else {}
    results["samples"]["run_calculations"] = {"status_code": run_resp.status_code, "body": run_json}
    check("run calculations endpoint succeeds", run_resp.status_code == 200 and run_json.get("processed", 0) > 0, results["samples"]["run_calculations"])
    check("run calculations unmapped_count is zero", run_resp.status_code == 200 and run_json.get("unmapped_count") == 0, run_json)
    unmapped = api_get(sess, api, "/calculations/unmapped-summary", params={"report_month": "2026-04"}).json()
    results["samples"]["unmapped_summary"] = unmapped
    check("unmapped summary empty after recalc", unmapped == [], unmapped)

    # Commission ALL fallback: look through public API pages for a calc matched to sub_category ALL.
    fallback_sample = None
    for skip in range(0, 24000, 2000):
        page = api_get(sess, api, "/calculations", params={**params_apr, "limit": 2000, "skip": skip, "sort_by": "computed_at", "sort_dir": "desc"}).json().get("items", [])
        if not page:
            break
        for c in page:
            rule = ((c.get("breakdown") or {}).get("commission_rule") or {})
            if str(rule.get("matched_sub_category") or "").strip().lower() == "all" and c.get("commission_pct") is not None:
                fallback_sample = {
                    "sales_id": c.get("sales_id"),
                    "sub_category": (c.get("breakdown") or {}).get("sub_category"),
                    "master_category": (c.get("breakdown") or {}).get("master_category"),
                    "matched_sub_category": rule.get("matched_sub_category"),
                    "commission_pct": c.get("commission_pct"),
                }
                break
        if fallback_sample:
            break
    results["samples"]["all_fallback"] = fallback_sample
    check("commission ALL fallback evidenced by matched_sub_category=ALL", fallback_sample is not None, fallback_sample)

    # Export workbook: 31 client-requested headers and joined calculation fields populated.
    export_resp = sess.get(api + "/sales/export", params=params_apr, timeout=180)
    export_ok = export_resp.status_code == 200 and export_resp.content.startswith(b"PK")
    export_detail = {"status_code": export_resp.status_code, "content_type": export_resp.headers.get("content-type"), "content_disposition": export_resp.headers.get("content-disposition"), "bytes": len(export_resp.content)}
    if export_ok:
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content), read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        data_rows = list(ws.iter_rows(min_row=2, max_row=21, values_only=True))
        header_index = {h: i for i, h in enumerate(headers)}
        required_headers = [
            "Brand", "Sale Type", "Posting Date", "Item No (SKU)", "Posting_Location Code", "Main Ctg", "Level No",
            "Price Range - Key (NSV)", "Price Range - Key (NSV after GT)", "Commission %", "Commission (Expected)",
            "GT Charge (Expected)", "Fixed Fee (Expected)", "Return Fee (Expected)", "Total Deductions", "Expected Settlement",
        ]
        populated_join_rows = 0
        for row in data_rows:
            if not row:
                continue
            if all(row[header_index[h]] not in (None, "") for h in ["Level No", "Price Range - Key (NSV)", "Price Range - Key (NSV after GT)"]):
                populated_join_rows += 1
        export_detail.update({
            "sheet": ws.title,
            "max_row": ws.max_row,
            "headers": headers,
            "header_count": len(headers),
            "first_20_join_populated_count": populated_join_rows,
        })
        check("sales export has 31 headers incl requested fields", len(headers) == 31 and all(h in headers for h in required_headers), export_detail)
        check("sales export has at least 20 data rows", ws.max_row >= 21, {"max_row": ws.max_row})
        check("sales export first 20 joined Level/Price Range fields populated", populated_join_rows == len(data_rows) and len(data_rows) >= 20, {"first_20_rows": len(data_rows), "populated_join_rows": populated_join_rows})
    else:
        check("sales export returns xlsx", False, export_detail)
    results["samples"]["export"] = export_detail

    OUT.write_text(json.dumps(results, indent=2, default=str))
    print(f"Wrote {OUT}")
    return 1 if results["failures"] else 0


if __name__ == "__main__":
    sys.exit(main())