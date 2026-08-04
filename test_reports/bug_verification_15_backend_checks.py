#!/usr/bin/env python3
"""Focused verification for April 2026 Myntra Sales Ledger/calculation bug."""

import json
import os
import re
from io import BytesIO
from pathlib import Path

import openpyxl
import requests
from pymongo import MongoClient


ROOT = Path("/app")
FRONTEND_ENV = ROOT / "frontend" / ".env"
BACKEND_ENV = ROOT / "backend" / ".env"
OUT = ROOT / "test_reports" / "bug_verification_15_backend_results.json"


def read_env(path):
    vals = {}
    if not path.exists():
        return vals
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        vals[k] = v.strip().strip('"').strip("'")
    return vals


def clean_loc(v):
    return isinstance(v, str) and bool(re.fullmatch(r"[A-Z0-9]+", v.strip()))


def invoice_prefix(inv):
    m = re.match(r"^([A-Za-z0-9]+)", (inv or "").strip())
    return m.group(1).upper() if m else None


def api_get(session, api_base, path, **params):
    r = session.get(f"{api_base}{path}", params={k: v for k, v in params.items() if v is not None}, timeout=120)
    r.raise_for_status()
    return r


def login(session, api_base):
    last = None
    for email in ("admin@fundle.ai", "admin@kazo.com"):
        r = session.post(f"{api_base}/auth/login", json={"email": email, "password": "admin123"}, timeout=60)
        last = {"email": email, "status_code": r.status_code, "body": r.text[:300]}
        if r.ok:
            data = r.json()
            session.headers.update({"Authorization": f"Bearer {data['token']}"})
            return {"email": email, "status_code": r.status_code, "user": data.get("user")}
    raise RuntimeError(f"Login failed: {last}")


def main():
    fe = read_env(FRONTEND_ENV)
    be = read_env(BACKEND_ENV)
    api_base = fe["REACT_APP_BACKEND_URL"].rstrip("/") + "/api"

    mongo = MongoClient(be["MONGO_URL"])
    db = mongo[be["DB_NAME"]]
    q = {"report_month": "2026-04", "portal": "myntra"}
    rows = list(db.sales.find(q, {"_id": 0, "id": 1, "txn_type": 1, "order_status": 1, "sales_invoice_no": 1, "posting_location_code": 1, "qty": 1}))

    db_checks = {
        "sales_row_count": len(rows),
        "missing_location_count": 0,
        "dirty_location_count": 0,
        "invoice_prefix_mismatch_count": 0,
        "location_counts": {},
        "sales_locations": {},
        "return_locations": {},
        "sample_bad_locations": [],
        "net_qty_sum": round(sum(float(r.get("qty") or 0) for r in rows), 2),
    }
    for r in rows:
        loc = r.get("posting_location_code")
        loc_s = (loc or "").strip() if isinstance(loc, str) else loc
        db_checks["location_counts"][loc_s or "<blank>"] = db_checks["location_counts"].get(loc_s or "<blank>", 0) + 1
        tx = (r.get("txn_type") or "").strip().lower()
        bucket = db_checks["return_locations"] if tx == "return" else db_checks["sales_locations"]
        bucket[loc_s or "<blank>"] = bucket.get(loc_s or "<blank>", 0) + 1
        if not loc_s:
            db_checks["missing_location_count"] += 1
            db_checks["sample_bad_locations"].append({"id": r.get("id"), "invoice": r.get("sales_invoice_no"), "loc": loc})
        elif not clean_loc(loc_s):
            db_checks["dirty_location_count"] += 1
            db_checks["sample_bad_locations"].append({"id": r.get("id"), "invoice": r.get("sales_invoice_no"), "loc": loc})
        pref = invoice_prefix(r.get("sales_invoice_no"))
        if pref and loc_s and pref != loc_s:
            db_checks["invoice_prefix_mismatch_count"] += 1
    db_checks["sample_bad_locations"] = db_checks["sample_bad_locations"][:10]

    session = requests.Session()
    api_checks = {}
    api_checks["login"] = login(session, api_base)
    summary = api_get(session, api_base, "/sales/summary", period_type="month", period_value="2026-04", portal="myntra").json()
    api_checks["sales_summary"] = summary

    sales_page = api_get(session, api_base, "/sales", period_type="month", period_value="2026-04", portal="myntra", txn_type="Sales", sort_by="order_date", sort_dir="desc", limit=20).json()
    return_page = api_get(session, api_base, "/sales", period_type="month", period_value="2026-04", portal="myntra", txn_type="Return", sort_by="order_date", sort_dir="desc", limit=20).json()
    api_checks["first_20_sales_locations"] = [x.get("posting_location_code") for x in sales_page["items"]]
    api_checks["first_20_return_locations"] = [x.get("posting_location_code") for x in return_page["items"]]
    api_checks["first_20_sales_all_myn"] = all(x.get("posting_location_code") == "MYN" for x in sales_page["items"])
    api_checks["first_20_return_all_clean_nonblank"] = all(clean_loc(x.get("posting_location_code")) for x in return_page["items"])
    api_checks["first_20_return_all_mysri"] = all(x.get("posting_location_code") == "MYSRI" for x in return_page["items"])

    # Re-run calculations for the exact April/Myntra period, matching the user action.
    run = session.post(f"{api_base}/calculations/run", json={"report_month": "2026-04", "portal": "myntra", "recalculate": True}, timeout=300)
    run.raise_for_status()
    api_checks["recalculate_result"] = run.json()

    calc_q = {"report_month": "2026-04", "portal": "myntra"}
    calc_count = db.calculations.count_documents(calc_q)
    calc_unmapped = db.calculations.count_documents({**calc_q, "unmapped": True})
    nonzero_taxes = db.calculations.count_documents({**calc_q, "$or": [{"tcs": {"$ne": 0}}, {"tds": {"$ne": 0}}, {"commission_gst": {"$ne": 0}}, {"fixed_fee_gst": {"$ne": 0}}]})
    return_dto_bad = db.calculations.count_documents({**calc_q, "order_type": "return_dto", "$or": [{"commission_incl_gst": {"$gt": 0}}, {"gt_charge": {"$gt": 0}}, {"fixed_fee_incl_gst": {"$gt": 0}}, {"return_fee": {"$lt": 0}}]})
    return_dto_sample = list(db.calculations.find({**calc_q, "order_type": "return_dto"}, {"_id": 0, "sales_id": 1, "commission_incl_gst": 1, "gt_charge": 1, "fixed_fee_incl_gst": 1, "return_fee": 1, "commission_gst": 1, "tcs": 1, "tds": 1}).limit(5))
    all_fallback_count = db.calculations.count_documents({**calc_q, "breakdown.master_category": "APPAREL", "breakdown.commission_rule.matched_sub_category": {"$regex": "^ALL$", "$options": "i"}, "unmapped": {"$ne": True}})
    api_checks["calculation_db"] = {
        "calc_count": calc_count,
        "unmapped_count_after_recalc": calc_unmapped,
        "nonzero_tcs_tds_or_gst_count": nonzero_taxes,
        "return_dto_bad_sign_count": return_dto_bad,
        "return_dto_sample": return_dto_sample,
        "apparel_all_commission_fallback_count": all_fallback_count,
    }

    # Excel export content and headers.
    export_resp = api_get(session, api_base, "/sales/export", period_type="month", period_value="2026-04", portal="myntra")
    wb = openpyxl.load_workbook(BytesIO(export_resp.content), read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    loc_idx = headers.index("Posting_Location Code") + 1 if "Posting_Location Code" in headers else None
    export_checks = {
        "status_code": export_resp.status_code,
        "content_type": export_resp.headers.get("content-type"),
        "sheet": ws.title,
        "row_count_excluding_header": ws.max_row - 1,
        "headers": headers,
        "required_header_presence": {},
        "location_blank_count": None,
        "location_dirty_count": None,
        "location_counts_sample": {},
    }
    required_exact = ["Brand", "Sale Type", "Posting Date", "Item No", "Posting_Location Code", "Main Ctg", "Level No", "Price Range Key (NSV)", "Price Range Key (NSV after GT)"]
    for h in required_exact:
        export_checks["required_header_presence"][h] = h in headers
    if loc_idx:
        blanks = dirty = 0
        counts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            loc = row[loc_idx - 1]
            loc_s = str(loc).strip() if loc is not None else ""
            counts[loc_s or "<blank>"] = counts.get(loc_s or "<blank>", 0) + 1
            if not loc_s:
                blanks += 1
            elif not clean_loc(loc_s):
                dirty += 1
        export_checks["location_blank_count"] = blanks
        export_checks["location_dirty_count"] = dirty
        export_checks["location_counts_sample"] = dict(sorted(counts.items())[:20])

    result = {
        "api_base": api_base,
        "db_checks": db_checks,
        "api_checks": api_checks,
        "export_checks": export_checks,
    }
    OUT.write_text(json.dumps(result, indent=2, default=str))
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()