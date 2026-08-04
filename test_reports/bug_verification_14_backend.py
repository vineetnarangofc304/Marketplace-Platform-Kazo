import io
import json
import os
from pathlib import Path

import openpyxl
import requests
from pymongo import MongoClient


BACKEND_URL = "https://marketplace-recon-1.preview.emergentagent.com".rstrip("/")
ADMIN_EMAIL = "admin@fundle.ai"
ADMIN_PASSWORD = "admin123"


def load_backend_env():
    env = {}
    p = Path("/app/backend/.env")
    if p.exists():
        for line in p.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k] = v.strip().strip('"').strip("'")
    return env


def assert_true(condition, message):
    if not condition:
        raise AssertionError(message)


def main():
    results = {"checks": [], "failures": []}

    def record(name, ok, details=None):
        entry = {"name": name, "ok": bool(ok), "details": details or {}}
        results["checks"].append(entry)
        print(json.dumps(entry, default=str))
        if not ok:
            results["failures"].append(entry)

    session = requests.Session()
    session.headers.update({"Content-Type": "application/json"})

    try:
        r = session.post(
            f"{BACKEND_URL}/api/auth/login",
            json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
            timeout=30,
        )
        assert_true(r.status_code == 200, f"login failed: {r.status_code} {r.text[:500]}")
        token = r.json()["token"]
        session.headers.update({"Authorization": f"Bearer {token}"})
        record("auth_login", True, {"email": ADMIN_EMAIL})
    except Exception as e:
        record("auth_login", False, {"error": str(e)})
        print(json.dumps(results, indent=2, default=str))
        raise

    # API proof for the exact regression: first 100 Myntra rows must expose MYN.
    r = session.get(f"{BACKEND_URL}/api/sales", params={"portal": "myntra", "limit": 100}, timeout=60)
    assert_true(r.status_code == 200, f"/api/sales failed: {r.status_code} {r.text[:500]}")
    sales = r.json()
    items = sales.get("items", [])
    bad_locations = [
        {"id": x.get("id"), "invoice": x.get("sales_invoice_no"), "posting_location_code": x.get("posting_location_code")}
        for x in items
        if x.get("posting_location_code") != "MYN"
    ]
    record(
        "api_sales_first_100_posting_location_code_myn",
        len(items) == 100 and not bad_locations,
        {"total": sales.get("total"), "sample_size": len(items), "bad_locations": bad_locations[:5]},
    )

    # DB proof for all existing Myntra rows, not just the API page.
    env = load_backend_env()
    client = MongoClient(env.get("MONGO_URL", "mongodb://localhost:27017"))
    db = client[env.get("DB_NAME", "test_database")]
    total_myntra_sales = db.sales.count_documents({"portal": "myntra"})
    non_myn = db.sales.count_documents({"portal": "myntra", "posting_location_code": {"$ne": "MYN"}})
    missing_loc = db.sales.count_documents({
        "portal": "myntra",
        "$or": [
            {"posting_location_code": {"$exists": False}},
            {"posting_location_code": None},
            {"posting_location_code": ""},
        ],
    })
    record(
        "db_all_myntra_sales_have_myn_location",
        total_myntra_sales == 21614 and non_myn == 0 and missing_loc == 0,
        {"total_myntra_sales": total_myntra_sales, "non_myn_count": non_myn, "missing_count": missing_loc},
    )

    # Sales summary: Order Qty (net) should be Sales rows - Return rows = 6,824.
    summary_params = {"period_type": "month", "period_value": "2026-04", "portal": "myntra"}
    r = session.get(f"{BACKEND_URL}/api/sales/summary", params=summary_params, timeout=60)
    assert_true(r.status_code == 200, f"/api/sales/summary failed: {r.status_code} {r.text[:500]}")
    summary = r.json()
    record(
        "sales_summary_net_order_qty_6824",
        summary.get("net_orders") == 6824,
        summary,
    )

    # Recalculate April Myntra and verify fully mapped.
    r = session.post(
        f"{BACKEND_URL}/api/calculations/run",
        json={"report_month": "2026-04", "portal": "myntra", "recalculate": True},
        timeout=300,
    )
    assert_true(r.status_code == 200, f"/api/calculations/run failed: {r.status_code} {r.text[:500]}")
    run = r.json()
    record(
        "recalculate_april_myntra_unmapped_zero",
        run.get("total_sales") == 21614 and run.get("unmapped_count") == 0,
        run,
    )
    r = session.get(
        f"{BACKEND_URL}/api/calculations",
        params={"period_type": "month", "period_value": "2026-04", "portal": "myntra", "severity_flag": "unmapped", "limit": 1},
        timeout=60,
    )
    assert_true(r.status_code == 200, f"/api/calculations unmapped failed: {r.status_code} {r.text[:500]}")
    unmapped = r.json()
    record("api_unmapped_total_zero_after_recalc", unmapped.get("total") == 0, {"total": unmapped.get("total")})

    # Return DTO calculation arithmetic and taxes.
    r = session.get(
        f"{BACKEND_URL}/api/calculations",
        params={"period_type": "month", "period_value": "2026-04", "portal": "myntra", "order_type": "return_dto", "limit": 20},
        timeout=60,
    )
    assert_true(r.status_code == 200, f"/api/calculations return_dto failed: {r.status_code} {r.text[:500]}")
    dto_rows = r.json().get("items", [])
    dto_bad = []
    for c in dto_rows:
        if not (
            c.get("commission_incl_gst") is not None and c.get("commission_incl_gst") < 0
            and c.get("gt_charge") is not None and c.get("gt_charge") < 0
            and c.get("fixed_fee_incl_gst") is not None and c.get("fixed_fee_incl_gst") < 0
            and c.get("return_fee") is not None and c.get("return_fee") > 0
            and c.get("tcs") == 0 and c.get("tds") == 0 and c.get("commission_gst") == 0
        ):
            dto_bad.append({
                "sales_id": c.get("sales_id"),
                "commission_incl_gst": c.get("commission_incl_gst"),
                "gt_charge": c.get("gt_charge"),
                "fixed_fee_incl_gst": c.get("fixed_fee_incl_gst"),
                "return_fee": c.get("return_fee"),
                "tcs": c.get("tcs"),
                "tds": c.get("tds"),
                "commission_gst": c.get("commission_gst"),
            })
    record(
        "return_dto_negative_reversals_return_fee_positive_taxes_zero",
        len(dto_rows) == 20 and not dto_bad,
        {"sample_size": len(dto_rows), "bad_rows": dto_bad[:5]},
    )

    # Sales calculations have TCS/TDS removed/zero.
    r = session.get(
        f"{BACKEND_URL}/api/calculations",
        params={"period_type": "month", "period_value": "2026-04", "portal": "myntra", "order_type": "sales", "limit": 100},
        timeout=60,
    )
    assert_true(r.status_code == 200, f"/api/calculations sales failed: {r.status_code} {r.text[:500]}")
    sales_calc_rows = r.json().get("items", [])
    tax_bad = [c.get("sales_id") for c in sales_calc_rows if c.get("tcs") != 0 or c.get("tds") != 0 or c.get("commission_gst") != 0]
    record(
        "sales_calc_tcs_tds_commission_gst_zero",
        len(sales_calc_rows) == 100 and not tax_bad,
        {"sample_size": len(sales_calc_rows), "bad_sales_ids": tax_bad[:5]},
    )

    # Apparels commission ALL fallback runtime evidence: at least some rows matched ALL and are fully mapped.
    all_fallback_count = db.calculations.count_documents({
        "portal": "myntra",
        "report_month": "2026-04",
        "breakdown.master_category": "APPAREL",
        "breakdown.commission_rule.matched_sub_category": {"$regex": "^ALL$", "$options": "i"},
        "unmapped": False,
    })
    record(
        "apparel_commission_all_fallback_present_and_mapped",
        all_fallback_count > 0,
        {"mapped_apparel_all_fallback_count": all_fallback_count},
    )

    # Export xlsx headers and Posting_Location Code values.
    r = session.get(f"{BACKEND_URL}/api/sales/export", params=summary_params, timeout=180)
    assert_true(r.status_code == 200, f"/api/sales/export failed: {r.status_code} {r.text[:500]}")
    content_type = r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required_headers = [
        "Brand", "Sale Type", "Posting Date", "Posting_Location Code", "Main Ctg", "Level No",
        "Price Range - Key (NSV)", "Price Range - Key (NSV after GT)",
    ]
    missing_headers = [h for h in required_headers if h not in headers]
    item_header_present = "Item No" in headers or "Item No (SKU)" in headers
    loc_idx = headers.index("Posting_Location Code") + 1 if "Posting_Location Code" in headers else None
    first_20_locations = []
    if loc_idx:
        for row in ws.iter_rows(min_row=2, max_row=21, values_only=True):
            first_20_locations.append(row[loc_idx - 1])
    record(
        "sales_export_xlsx_headers_and_first_20_locations",
        "spreadsheet" in content_type and not missing_headers and item_header_present and len(first_20_locations) == 20 and all(v == "MYN" for v in first_20_locations),
        {
            "content_type": content_type,
            "headers": headers,
            "missing_headers": missing_headers,
            "item_header_present": item_header_present,
            "first_20_locations": first_20_locations,
        },
    )

    out = Path("/app/test_reports/bug_verification_14_backend_results.json")
    out.write_text(json.dumps(results, indent=2, default=str))
    print(f"WROTE {out}")
    if results["failures"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()