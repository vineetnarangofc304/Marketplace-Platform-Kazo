"""KAZO Marketplace — Monthly reports, masters (settlement/tax), and unmapped-order tests."""
import os
import requests
import pytest

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://settlement-intel-1.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "admin@fundle.ai"
ADMIN_PASSWORD = "admin123"
MONTH = "2026-04"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    s.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return s


# ---------- Reports ----------
def test_reports_months_includes_apr26(auth):
    r = auth.get(f"{BASE_URL}/api/reports/months")
    assert r.status_code == 200
    months = r.json()
    assert isinstance(months, list)
    assert MONTH in months, f"Expected {MONTH} in months, got {months}"


def test_reports_monthly(auth):
    r = auth.get(f"{BASE_URL}/api/reports/monthly", params={"month": MONTH})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["month"] == MONTH
    assert "kpi" in data and "by_category" in data and "by_sub_category" in data and "reconciliation" in data
    kpi = data["kpi"]
    # After data is uploaded / calculated: KPI values should be non-zero
    assert kpi.get("total_orders", 0) > 0, f"KPI has no orders: {kpi}"
    assert kpi.get("total_nsv") or kpi.get("sales_nsv"), f"KPI NSV is zero: {kpi}"
    assert kpi.get("expected_commission", 0) > 0
    assert isinstance(data["by_category"], list) and len(data["by_category"]) > 0
    assert isinstance(data["by_sub_category"], list) and len(data["by_sub_category"]) > 0


def test_reports_monthly_bad_month(auth):
    r = auth.get(f"{BASE_URL}/api/reports/monthly", params={"month": "bad"})
    assert r.status_code == 400


def test_reports_monthly_export_xlsx(auth):
    r = auth.get(f"{BASE_URL}/api/reports/monthly/export", params={"month": MONTH})
    assert r.status_code == 200, r.text[:400]
    ct = r.headers.get("content-type", "")
    assert "openxmlformats-officedocument.spreadsheetml.sheet" in ct, f"Wrong content-type: {ct}"
    assert len(r.content) > 5000, "xlsx body suspiciously small"
    # Parse xlsx to verify sheet names
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    for want in ["Summary", "By Category", "By Sub-Category", "By Zone", "Order Detail", "Discrepancies", "Unmapped Orders"]:
        assert want in wb.sheetnames, f"Missing sheet {want}. Have: {wb.sheetnames}"


# ---------- Unmapped calculations ----------
def test_unmapped_summary(auth):
    r = auth.get(f"{BASE_URL}/api/calculations/unmapped-summary")
    assert r.status_code == 200, r.text
    body = r.json()
    assert isinstance(body, (list, dict))


def test_calculations_unmapped_only_filter(auth):
    r = auth.get(f"{BASE_URL}/api/calculations", params={"unmapped_only": "true", "limit": 5})
    assert r.status_code == 200
    body = r.json()
    assert "total" in body
    # All returned items must have unmapped=True
    for it in body.get("items", []):
        assert it.get("unmapped") is True, f"Non-unmapped item leaked: {it.get('id')}"


def test_calculations_report_month_filter(auth):
    r = auth.get(f"{BASE_URL}/api/calculations", params={"report_month": MONTH, "limit": 5})
    assert r.status_code == 200
    body = r.json()
    for it in body.get("items", []):
        assert it.get("report_month") == MONTH


def test_unmapped_count_is_small(auth):
    """After full seed, unmapped ratio should be < 1%."""
    total = auth.get(f"{BASE_URL}/api/calculations?limit=1").json()["total"]
    unmapped = auth.get(f"{BASE_URL}/api/calculations?unmapped_only=true&limit=1").json()["total"]
    assert total > 0
    ratio = unmapped / total
    assert ratio < 0.01, f"Too many unmapped: {unmapped}/{total} = {ratio:.2%}"


# ---------- Sales months ----------
def test_sales_months(auth):
    r = auth.get(f"{BASE_URL}/api/sales/months")
    assert r.status_code == 200
    months = r.json()
    # Endpoint may return list[str] or list[{month, count}]
    flat = [m["month"] if isinstance(m, dict) else m for m in months]
    assert MONTH in flat, f"Expected {MONTH} in sales months, got {months}"


# ---------- Masters: settlement settings & tax rates ----------
def test_settlement_settings_get_and_set(auth):
    r = auth.get(f"{BASE_URL}/api/masters/settlement-settings")
    assert r.status_code == 200
    original = r.json()
    # Toggle values
    payload = {
        "default_zone_when_missing": "Local",
        "treat_dash_as_missing_zone": False,
        "apply_default_zone": False,
    }
    r2 = auth.post(f"{BASE_URL}/api/masters/settlement-settings", json=payload)
    assert r2.status_code == 200
    got = r2.json()
    assert got["default_zone_when_missing"] == "Local"
    assert got["treat_dash_as_missing_zone"] is False
    assert got["apply_default_zone"] is False
    # Verify persistence
    r3 = auth.get(f"{BASE_URL}/api/masters/settlement-settings")
    assert r3.json()["default_zone_when_missing"] == "Local"
    # Revert
    auth.post(f"{BASE_URL}/api/masters/settlement-settings", json={
        "default_zone_when_missing": original.get("default_zone_when_missing", "Zonal"),
        "treat_dash_as_missing_zone": original.get("treat_dash_as_missing_zone", True),
        "apply_default_zone": original.get("apply_default_zone", True),
    })


def test_tax_rates_get_and_set(auth):
    r = auth.get(f"{BASE_URL}/api/masters/tax-rates")
    assert r.status_code == 200
    orig = r.json()
    assert "gst_rate" in orig and "tcs_rate" in orig and "tds_rate" in orig
    r2 = auth.post(f"{BASE_URL}/api/masters/tax-rates", json={
        "gst_rate": 0.18, "tcs_rate": 0.01, "tds_rate": 0.001
    })
    assert r2.status_code == 200
    assert abs(r2.json()["tcs_rate"] - 0.01) < 1e-9
    # Revert
    auth.post(f"{BASE_URL}/api/masters/tax-rates", json={
        "gst_rate": orig.get("gst_rate", 0.18),
        "tcs_rate": orig.get("tcs_rate", 0.005),
        "tds_rate": orig.get("tds_rate", 0.001),
    })


def test_commission_rules_count(auth):
    r = auth.get(f"{BASE_URL}/api/masters/commission-rules")
    assert r.status_code == 200
    rules = r.json()
    assert len(rules) >= 170, f"Expected >=170 rules, got {len(rules)}"


def test_masters_seed_counts(auth):
    fixed = auth.get(f"{BASE_URL}/api/masters/fixed-fees").json()
    assert len(fixed) == 6, f"Fixed fees count: {len(fixed)}"
    gt = auth.get(f"{BASE_URL}/api/masters/gt-charges").json()
    assert len(gt) >= 260, f"GT charges count: {len(gt)}"
    ret = auth.get(f"{BASE_URL}/api/masters/return-fees").json()
    assert len(ret) == 15, f"Return fees: {len(ret)}"
    lvls = auth.get(f"{BASE_URL}/api/masters/subcat-levels").json()
    assert len(lvls) >= 45, f"Subcat levels: {len(lvls)}"


def test_reconciliation_with_report_month(auth):
    r = auth.post(f"{BASE_URL}/api/reconciliation/run", json={"report_month": MONTH})
    assert r.status_code == 200, r.text
    body = r.json()
    assert "matched" in body and "variance" in body
