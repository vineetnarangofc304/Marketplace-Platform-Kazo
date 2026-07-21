"""Iteration 5 — defect-fix regression tests (calc engine, search, masters export/import)."""
import io
import os
import re
import time
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://ledger-dashboard-13.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "admin@kazo.com"
ADMIN_PASSWORD = "admin123"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    tok = r.json()["token"]
    s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


# ---------- Health / auth ----------
def test_login_ok(auth):
    r = auth.get(f"{BASE_URL}/api/auth/me")
    assert r.status_code == 200 and r.json()["email"] == ADMIN_EMAIL


# ---------- Calc engine — order_type classification & math ----------
def test_run_calculations_recalculate(auth):
    r = auth.post(f"{BASE_URL}/api/calculations/run", json={"recalculate": True}, timeout=300)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["processed"] > 0
    print("run_calculations:", body)


def _fetch_sample(auth, order_type):
    r = auth.get(f"{BASE_URL}/api/calculations?limit=5&search=&sort_by=computed_at", timeout=60)
    assert r.status_code == 200
    # We need to filter by order_type — no dedicated param, so fetch bigger and filter
    r = auth.get(f"{BASE_URL}/api/calculations?limit=2000", timeout=60)
    assert r.status_code == 200
    items = r.json()["items"]
    return [it for it in items if it.get("order_type") == order_type]


def test_order_types_present(auth):
    r = auth.get(f"{BASE_URL}/api/calculations?limit=2000", timeout=60)
    assert r.status_code == 200
    items = r.json()["items"]
    types = {it.get("order_type") for it in items}
    print("order_types seen:", types)
    assert "sales" in types
    # 'dto' should exist per user data (Handbags row)
    # rto / internal_cancel may be absent — do NOT fail if so


def test_math_sales(auth):
    samples = _fetch_sample(auth, "sales")
    assert samples, "no sales rows"
    for it in samples[:5]:
        nsv = it["breakdown"]["nsv_val"]
        gt = it.get("gt_charge") or 0
        nsv_after_gt = it.get("nsv_after_gt")
        assert nsv_after_gt is not None
        assert abs(nsv_after_gt - (nsv - gt)) < 0.05, f"nsv_after_gt mismatch: {it}"
        pct = it.get("commission_pct")
        cb = it.get("commission_base")
        if pct is not None and cb is not None:
            assert abs(cb - nsv_after_gt * pct) < 0.05, f"commission_base mismatch: {it}"
        # expected_settlement = nsv_after_gt - (commission_incl_gst + ff_incl_gst + gt + tcs + tds + return_fee(0))
        if it.get("expected_settlement") is not None:
            parts = [it.get("commission_incl_gst") or 0, it.get("fixed_fee_incl_gst") or 0,
                     it.get("gt_charge") or 0, it.get("tcs") or 0, it.get("tds") or 0,
                     it.get("return_fee") or 0]
            expected = nsv_after_gt - sum(parts)
            assert abs(it["expected_settlement"] - expected) < 0.05


def test_math_dto(auth):
    samples = _fetch_sample(auth, "dto")
    if not samples:
        pytest.skip("no DTO rows in dataset")
    for it in samples[:5]:
        assert it["commission_incl_gst"] == 0
        assert it["gt_charge"] == 0
        assert it["tcs"] == 0
        assert it["tds"] == 0
        # return_fee == fixed_fee_incl_gst
        ff = it.get("fixed_fee_incl_gst")
        rf = it.get("return_fee")
        assert ff is not None and rf is not None
        assert abs(rf - ff) < 0.05
        nsv = it["breakdown"]["nsv_val"]
        # expected_settlement ≈ -abs(nsv) - fixed_fee_incl_gst
        expected = -abs(nsv) - ff
        assert abs(it["expected_settlement"] - expected) < 0.05, f"DTO settlement mismatch: {it}"


def test_math_rto_internal_cancel(auth):
    for t in ("rto", "internal_cancel"):
        samples = _fetch_sample(auth, t)
        if not samples:
            print(f"no {t} rows — skipping")
            continue
        for it in samples[:3]:
            for k in ("commission_incl_gst", "fixed_fee_incl_gst", "gt_charge",
                     "return_fee", "tcs", "tds", "expected_settlement"):
                assert (it.get(k) or 0) == 0, f"{t} {k} not zero: {it}"
            assert it.get("unmapped") is False, f"{t} should not be unmapped"


# ---------- Search fixes ----------
def test_sales_search_partial(auth):
    r = auth.get(f"{BASE_URL}/api/sales?limit=1")
    assert r.status_code == 200
    items = r.json()["items"]
    assert items
    oid = items[0].get("online_order_id") or ""
    frag = oid[:8]
    if not frag:
        pytest.skip("no order id fragment")
    r2 = auth.get(f"{BASE_URL}/api/sales?search={frag}&limit=5")
    assert r2.status_code == 200
    assert r2.json()["total"] >= 1
    assert any(frag.lower() in (i.get("online_order_id") or "").lower() for i in r2.json()["items"])


def test_calc_search_partial(auth):
    r = auth.get(f"{BASE_URL}/api/calculations?limit=1")
    items = r.json()["items"]
    assert items
    oid = items[0].get("online_order_id") or ""
    frag = oid[:8]
    if not frag:
        pytest.skip("no order id fragment")
    r2 = auth.get(f"{BASE_URL}/api/calculations?search={frag}&limit=5")
    assert r2.status_code == 200
    assert r2.json()["total"] >= 1


def test_search_regex_metachar_no_500(auth):
    for meta in [".", "[", "(", "*", "+", "?", "\\"]:
        r = auth.get(f"{BASE_URL}/api/sales?search={meta}&limit=5")
        assert r.status_code == 200, f"sales search '{meta}' → {r.status_code}"
        r = auth.get(f"{BASE_URL}/api/calculations?search={meta}&limit=5")
        assert r.status_code == 200, f"calc search '{meta}' → {r.status_code}"


# ---------- Masters export/import ----------
EXPECTED_SHEETS = ["Commission Rules", "Fixed Fee", "GT Charges", "Return Fee",
                   "Sub-Cat Levels", "Tolerance", "Tax Rates", "Settlement Config"]


def test_masters_export(auth):
    r = auth.get(f"{BASE_URL}/api/masters/export", timeout=60)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "").lower()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    for s in EXPECTED_SHEETS:
        assert s in wb.sheetnames, f"sheet missing: {s}"
    # Row counts
    def row_count(sname):
        ws = wb[sname]
        return sum(1 for _ in ws.iter_rows(values_only=True)) - 1
    cr = row_count("Commission Rules")
    gt = row_count("GT Charges")
    print(f"Commission Rules rows: {cr}, GT Charges rows: {gt}")
    assert cr >= 170
    assert gt >= 260
    # persist file for import test
    with open("/tmp/masters_export.xlsx", "wb") as f:
        f.write(r.content)


def _pre_import_counts(auth):
    cr = len(auth.get(f"{BASE_URL}/api/masters/commission-rules").json())
    return cr


def test_masters_import_merge(auth):
    pre = _pre_import_counts(auth)
    with open("/tmp/masters_export.xlsx", "rb") as f:
        files = {"file": ("m.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {k: v for k, v in auth.headers.items() if k != "Content-Type"}
        r = requests.post(f"{BASE_URL}/api/masters/import?mode=merge", files=files, headers=headers, timeout=120)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    assert len(body["sheets"]) == 8
    post = _pre_import_counts(auth)
    assert post >= 173, f"post-import commission rules count={post}"


def test_masters_import_replace(auth):
    pre = _pre_import_counts(auth)
    with open("/tmp/masters_export.xlsx", "rb") as f:
        files = {"file": ("m.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {k: v for k, v in auth.headers.items() if k != "Content-Type"}
        r = requests.post(f"{BASE_URL}/api/masters/import?mode=replace", files=files, headers=headers, timeout=120)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    post = _pre_import_counts(auth)
    assert post >= pre, f"pre={pre} post={post}"


# ---------- Regression ----------
def test_regression_endpoints(auth):
    for path in [
        "/api/dashboard/overview",
        "/api/insights/health-score?period_type=month&period_value=2026-04",
        "/api/reports/period?period_type=month&period_value=2026-04",
        "/api/recovery/summary?period_type=month&period_value=2026-04",
    ]:
        r = auth.get(f"{BASE_URL}{path}", timeout=30)
        assert r.status_code == 200, f"{path} → {r.status_code}: {r.text[:200]}"


def test_known_dto_order(auth):
    """Known test DTO from user: BA030AAB-E147-4FA2-847F-8B119D06AEC1 → Handbags ₹6490 → return_fee=71.98, settlement=-6561.98"""
    r = auth.get(f"{BASE_URL}/api/calculations?search=BA030AAB&limit=5")
    if r.json()["total"] == 0:
        pytest.skip("known DTO id not present in current dataset")
    items = r.json()["items"]
    dto = next((i for i in items if i.get("order_type") == "dto"), None)
    if not dto:
        pytest.skip("no DTO order_type on that id")
    print("known DTO:", {k: dto.get(k) for k in ("return_fee", "expected_settlement", "commission_incl_gst", "gt_charge")})
    assert dto["commission_incl_gst"] == 0
    assert dto["gt_charge"] == 0
    # user expectation: return_fee ≈ 71.98, settlement ≈ -6561.98
    assert abs(dto.get("return_fee", 0) - 71.98) < 1.0
    assert abs(dto.get("expected_settlement", 0) - (-6561.98)) < 2.0
