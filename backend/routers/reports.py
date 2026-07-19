"""Monthly Report generator — JSON aggregates + Excel export.

Aggregates per-month:
- Header KPIs (Total NSV, Expected Commission, Deductions, Settlement)
- Breakdown by category, sub-category, zone
- Reconciliation summary (if settlement rows exist for the month)
- Order-level detail sheet
- Discrepancy sheet
- Unmapped-orders sheet (for ops to fix masters)
"""
import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from db import db

router = APIRouter(tags=["reports"])


@router.get("/reports/months")
async def available_months():
    """Union of report_month values from sales, settlement, and discrepancies."""
    months = set()
    async for r in db.sales.aggregate([
        {"$match": {"report_month": {"$ne": None}}},
        {"$group": {"_id": "$report_month"}},
    ]):
        months.add(r["_id"])
    async for r in db.settlement.aggregate([
        {"$match": {"report_month": {"$ne": None}}},
        {"$group": {"_id": "$report_month"}},
    ]):
        months.add(r["_id"])
    return sorted(list(months))


async def _month_aggregate(month: str) -> Dict[str, Any]:
    # KPI (from calculations for this month)
    calc_match = {"report_month": month}
    kpi_pipe = [
        {"$match": calc_match},
        {"$group": {
            "_id": None,
            "total_orders": {"$sum": 1},
            "total_nsv": {"$sum": {"$ifNull": ["$breakdown.nsv_val", 0]}},
            "expected_commission": {"$sum": {"$ifNull": ["$commission_incl_gst", 0]}},
            "expected_fixed_fee": {"$sum": {"$ifNull": ["$fixed_fee_incl_gst", 0]}},
            "expected_gt_charge": {"$sum": {"$ifNull": ["$gt_charge", 0]}},
            "expected_return_fee": {"$sum": {"$ifNull": ["$return_fee", 0]}},
            "expected_tcs": {"$sum": {"$ifNull": ["$tcs", 0]}},
            "expected_tds": {"$sum": {"$ifNull": ["$tds", 0]}},
            "expected_deductions": {"$sum": {"$ifNull": ["$total_deductions", 0]}},
            "expected_settlement": {"$sum": {"$ifNull": ["$expected_settlement", 0]}},
            "unmapped_orders": {"$sum": {"$cond": ["$unmapped", 1, 0]}},
        }},
    ]
    kpi = await db.calculations.aggregate(kpi_pipe).to_list(1)
    kpi = kpi[0] if kpi else {}
    kpi.pop("_id", None)

    # Sales-only NSV (for sales rows not yet calculated)
    sales_kpi = await db.sales.aggregate([
        {"$match": {"report_month": month}},
        {"$group": {"_id": None, "sales_rows": {"$sum": 1}, "sales_nsv": {"$sum": "$nsv_val"}}},
    ]).to_list(1)
    if sales_kpi:
        sales_kpi[0].pop("_id", None)
        kpi.update(sales_kpi[0])

    # By category
    by_cat = await db.calculations.aggregate([
        {"$match": calc_match},
        {"$group": {
            "_id": "$breakdown.master_category",
            "orders": {"$sum": 1},
            "nsv": {"$sum": {"$ifNull": ["$breakdown.nsv_val", 0]}},
            "commission": {"$sum": {"$ifNull": ["$commission_incl_gst", 0]}},
            "fixed_fee": {"$sum": {"$ifNull": ["$fixed_fee_incl_gst", 0]}},
            "gt_charge": {"$sum": {"$ifNull": ["$gt_charge", 0]}},
            "expected_settlement": {"$sum": {"$ifNull": ["$expected_settlement", 0]}},
        }},
        {"$sort": {"nsv": -1}},
    ]).to_list(50)

    by_subcat = await db.calculations.aggregate([
        {"$match": calc_match},
        {"$group": {
            "_id": "$breakdown.sub_category",
            "orders": {"$sum": 1},
            "nsv": {"$sum": {"$ifNull": ["$breakdown.nsv_val", 0]}},
            "commission": {"$sum": {"$ifNull": ["$commission_incl_gst", 0]}},
            "fixed_fee": {"$sum": {"$ifNull": ["$fixed_fee_incl_gst", 0]}},
            "gt_charge": {"$sum": {"$ifNull": ["$gt_charge", 0]}},
            "expected_settlement": {"$sum": {"$ifNull": ["$expected_settlement", 0]}},
        }},
        {"$sort": {"nsv": -1}},
    ]).to_list(100)

    by_zone = await db.calculations.aggregate([
        {"$match": calc_match},
        {"$group": {
            "_id": "$breakdown.zone",
            "orders": {"$sum": 1},
            "nsv": {"$sum": {"$ifNull": ["$breakdown.nsv_val", 0]}},
            "gt_charge": {"$sum": {"$ifNull": ["$gt_charge", 0]}},
        }},
    ]).to_list(20)

    # Recon summary
    disc_match = {"report_month": month}
    total_disc = await db.discrepancies.count_documents(disc_match)
    disc_by_sev = await db.discrepancies.aggregate([
        {"$match": disc_match},
        {"$group": {"_id": "$severity", "count": {"$sum": 1}, "recoverable": {"$sum": "$recoverable"}}},
    ]).to_list(10)
    total_recoverable = await db.discrepancies.aggregate([
        {"$match": disc_match},
        {"$group": {"_id": None, "sum": {"$sum": "$recoverable"}}},
    ]).to_list(1)
    total_recoverable = total_recoverable[0]["sum"] if total_recoverable else 0

    return {
        "month": month,
        "kpi": kpi,
        "by_category": [{"category": c["_id"], **{k: v for k, v in c.items() if k != "_id"}} for c in by_cat],
        "by_sub_category": [{"sub_category": c["_id"], **{k: v for k, v in c.items() if k != "_id"}} for c in by_subcat],
        "by_zone": [{"zone": c["_id"], **{k: v for k, v in c.items() if k != "_id"}} for c in by_zone],
        "reconciliation": {
            "total_discrepancies": total_disc,
            "total_recoverable": round(total_recoverable or 0, 2),
            "by_severity": [
                {"severity": s["_id"], "count": s["count"], "recoverable": round(s.get("recoverable", 0), 2)}
                for s in disc_by_sev
            ],
        },
    }


@router.get("/reports/monthly")
async def monthly_report(month: str = Query(..., description="YYYY-MM")):
    if not month or len(month) != 7:
        raise HTTPException(400, "month must be YYYY-MM")
    return await _month_aggregate(month)


def _fmt_num(x):
    if x is None:
        return None
    try:
        return round(float(x), 2)
    except Exception:
        return x


@router.get("/reports/monthly/export")
async def export_monthly(month: str = Query(..., description="YYYY-MM")):
    """Return an Excel workbook with the full monthly report."""
    agg = await _month_aggregate(month)

    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    title_font = Font(name="Calibri", size=14, bold=True)
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    money_fmt = '#,##0.00'

    def write_header(ws, headers, row=1):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    def autosize(ws, max_col):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18

    # 1. Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(1, 1, "KAZO Marketplace Finance — Monthly Report").font = title_font
    ws.cell(2, 1, f"Marketplace: Myntra   |   Report Month: {month}   |   Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = subtitle_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    kpi = agg.get("kpi") or {}
    kpi_rows = [
        ("Sales Rows", kpi.get("sales_rows", 0)),
        ("Total Orders (calculated)", kpi.get("total_orders", 0)),
        ("Unmapped Orders", kpi.get("unmapped_orders", 0)),
        ("Total NSV (₹)", _fmt_num(kpi.get("total_nsv") or kpi.get("sales_nsv"))),
        ("Expected Commission (incl GST) (₹)", _fmt_num(kpi.get("expected_commission"))),
        ("Expected Fixed Fee (incl GST) (₹)", _fmt_num(kpi.get("expected_fixed_fee"))),
        ("Expected GT / Logistics (₹)", _fmt_num(kpi.get("expected_gt_charge"))),
        ("Expected Return Fee (₹)", _fmt_num(kpi.get("expected_return_fee"))),
        ("Expected TCS (₹)", _fmt_num(kpi.get("expected_tcs"))),
        ("Expected TDS (₹)", _fmt_num(kpi.get("expected_tds"))),
        ("Total Expected Deductions (₹)", _fmt_num(kpi.get("expected_deductions"))),
        ("Expected Net Settlement (₹)", _fmt_num(kpi.get("expected_settlement"))),
    ]
    write_header(ws, ["Metric", "Value"], row=4)
    for i, (k, v) in enumerate(kpi_rows, start=5):
        ws.cell(i, 1, k).font = Font(bold=True)
        c = ws.cell(i, 2, v)
        if isinstance(v, (int, float)):
            c.number_format = money_fmt
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    # Reconciliation KPIs
    recon = agg.get("reconciliation", {})
    ws.cell(len(kpi_rows) + 7, 1, "Reconciliation Summary").font = title_font
    write_header(ws, ["Metric", "Value"], row=len(kpi_rows) + 8)
    r = len(kpi_rows) + 9
    ws.cell(r, 1, "Total Discrepancies").font = Font(bold=True)
    ws.cell(r, 2, recon.get("total_discrepancies", 0))
    r += 1
    ws.cell(r, 1, "Total Recoverable (₹)").font = Font(bold=True)
    c = ws.cell(r, 2, recon.get("total_recoverable", 0))
    c.number_format = money_fmt
    for s in recon.get("by_severity", []):
        r += 1
        ws.cell(r, 1, f"— {s['severity'].upper()} ({s['count']} cases)").font = Font(bold=True)
        c = ws.cell(r, 2, s["recoverable"])
        c.number_format = money_fmt

    # 2. By Category
    ws2 = wb.create_sheet("By Category")
    headers = ["Master Category", "Orders", "NSV", "Commission (incl GST)", "Fixed Fee", "GT Charge", "Expected Settlement"]
    write_header(ws2, headers)
    for i, r in enumerate(agg["by_category"], start=2):
        ws2.cell(i, 1, r.get("category"))
        ws2.cell(i, 2, r.get("orders"))
        for col_i, k in enumerate(["nsv", "commission", "fixed_fee", "gt_charge", "expected_settlement"], start=3):
            c = ws2.cell(i, col_i, _fmt_num(r.get(k)))
            c.number_format = money_fmt
    autosize(ws2, len(headers))

    # 3. By Sub-Category
    ws3 = wb.create_sheet("By Sub-Category")
    headers = ["Sub Category", "Orders", "NSV", "Commission (incl GST)", "Fixed Fee", "GT Charge", "Expected Settlement"]
    write_header(ws3, headers)
    for i, r in enumerate(agg["by_sub_category"], start=2):
        ws3.cell(i, 1, r.get("sub_category"))
        ws3.cell(i, 2, r.get("orders"))
        for col_i, k in enumerate(["nsv", "commission", "fixed_fee", "gt_charge", "expected_settlement"], start=3):
            c = ws3.cell(i, col_i, _fmt_num(r.get(k)))
            c.number_format = money_fmt
    autosize(ws3, len(headers))

    # 4. By Zone
    ws4 = wb.create_sheet("By Zone")
    headers = ["Zone", "Orders", "NSV", "GT Charge"]
    write_header(ws4, headers)
    for i, r in enumerate(agg["by_zone"], start=2):
        ws4.cell(i, 1, r.get("zone"))
        ws4.cell(i, 2, r.get("orders"))
        c = ws4.cell(i, 3, _fmt_num(r.get("nsv"))); c.number_format = money_fmt
        c = ws4.cell(i, 4, _fmt_num(r.get("gt_charge"))); c.number_format = money_fmt
    autosize(ws4, len(headers))

    # 5. Order-level detail (calculations join sales)
    ws5 = wb.create_sheet("Order Detail")
    detail_headers = [
        "Order ID", "SKU", "Order Date", "Category", "Sub-Category", "Zone", "Qty",
        "MRP", "Customer Discount", "NSV", "Commission %", "Commission (incl GST)",
        "Fixed Fee (incl GST)", "GT Charge", "Return Fee", "TCS", "TDS",
        "Expected Deductions", "Expected Settlement", "Unmapped?",
    ]
    write_header(ws5, detail_headers)
    row_i = 2
    # Join in Python — pull calculations for this month, then sales by id
    calcs = await db.calculations.find({"report_month": month}, {"_id": 0}).to_list(100000)
    sale_ids = [c["sales_id"] for c in calcs]
    sales_by_id = {}
    async for s in db.sales.find({"id": {"$in": sale_ids}}, {"_id": 0}):
        sales_by_id[s["id"]] = s
    for c in calcs:
        s = sales_by_id.get(c["sales_id"], {})
        ws5.cell(row_i, 1, c.get("online_order_id"))
        ws5.cell(row_i, 2, c.get("sku"))
        ws5.cell(row_i, 3, (s.get("order_date") or "")[:10])
        ws5.cell(row_i, 4, s.get("category"))
        ws5.cell(row_i, 5, s.get("sub_category"))
        ws5.cell(row_i, 6, s.get("zone"))
        ws5.cell(row_i, 7, s.get("qty"))
        for col_i, key in enumerate([
            "mrp_from_sale", "customer_discount_from_sale", "nsv_val_from_sale",
        ], start=8):
            pass
        c8 = ws5.cell(row_i, 8, _fmt_num(s.get("mrp"))); c8.number_format = money_fmt
        c9 = ws5.cell(row_i, 9, _fmt_num(s.get("customer_discount"))); c9.number_format = money_fmt
        c10 = ws5.cell(row_i, 10, _fmt_num(s.get("nsv_val"))); c10.number_format = money_fmt
        c11 = ws5.cell(row_i, 11, _fmt_num((c.get("commission_pct") or 0) * 100)); c11.number_format = '0.00"%"'
        for col_i, key in enumerate([
            "commission_incl_gst", "fixed_fee_incl_gst", "gt_charge", "return_fee",
            "tcs", "tds", "total_deductions", "expected_settlement",
        ], start=12):
            cell = ws5.cell(row_i, col_i, _fmt_num(c.get(key)))
            cell.number_format = money_fmt
        ws5.cell(row_i, 20, "YES" if c.get("unmapped") else "no")
        row_i += 1
    autosize(ws5, len(detail_headers))
    ws5.freeze_panes = "A2"

    # 6. Discrepancies
    ws6 = wb.create_sheet("Discrepancies")
    d_headers = ["Order ID", "SKU", "Severity", "Match Status", "Reason", "Settle Variance", "Recoverable", "Expected Commission", "Actual Commission", "Expected GT", "Actual GT"]
    write_header(ws6, d_headers)
    row_i = 2
    async for d in db.discrepancies.find({"report_month": month}, {"_id": 0}):
        ws6.cell(row_i, 1, d.get("online_order_id"))
        ws6.cell(row_i, 2, d.get("sku"))
        ws6.cell(row_i, 3, (d.get("severity") or "").upper())
        ws6.cell(row_i, 4, d.get("match_status"))
        ws6.cell(row_i, 5, d.get("reason"))
        c = ws6.cell(row_i, 6, _fmt_num(d.get("settle_variance"))); c.number_format = money_fmt
        c = ws6.cell(row_i, 7, _fmt_num(d.get("recoverable"))); c.number_format = money_fmt
        # Pull component-level expected/actual
        comps = {x.get("component"): x for x in (d.get("components") or [])}
        c = ws6.cell(row_i, 8, _fmt_num((comps.get("commission") or {}).get("expected"))); c.number_format = money_fmt
        c = ws6.cell(row_i, 9, _fmt_num((comps.get("commission") or {}).get("actual"))); c.number_format = money_fmt
        c = ws6.cell(row_i, 10, _fmt_num((comps.get("gt_charge") or {}).get("expected"))); c.number_format = money_fmt
        c = ws6.cell(row_i, 11, _fmt_num((comps.get("gt_charge") or {}).get("actual"))); c.number_format = money_fmt
        row_i += 1
    autosize(ws6, len(d_headers))
    ws6.freeze_panes = "A2"

    # 7. Unmapped
    ws7 = wb.create_sheet("Unmapped Orders")
    u_headers = ["Order ID", "SKU", "Sub-Category", "Zone", "ISP", "Reasons"]
    write_header(ws7, u_headers)
    row_i = 2
    async for c in db.calculations.find({"report_month": month, "unmapped": True}, {"_id": 0}):
        ws7.cell(row_i, 1, c.get("online_order_id"))
        ws7.cell(row_i, 2, c.get("sku"))
        ws7.cell(row_i, 3, (c.get("breakdown") or {}).get("sub_category"))
        ws7.cell(row_i, 4, (c.get("breakdown") or {}).get("zone"))
        cell = ws7.cell(row_i, 5, _fmt_num((c.get("breakdown") or {}).get("isp"))); cell.number_format = money_fmt
        ws7.cell(row_i, 6, "; ".join(c.get("unmapped_reasons") or []))
        row_i += 1
    autosize(ws7, len(u_headers))
    ws7.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"KAZO_Myntra_Report_{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
