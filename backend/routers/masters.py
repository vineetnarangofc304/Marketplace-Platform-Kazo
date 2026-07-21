"""Commission masters: commission rules, fixed fee slabs, GT charges, return fees,
sub-category level mapping, and tolerance settings.

Rules are loaded from /app/backend/data_myntra_commission_rules.json (173 authoritative
Myntra rules extracted from the KAZO Myntra commission master file).
"""
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field

from db import db
from deps import require_admin

router = APIRouter(tags=["masters"])
ROOT_DIR = Path(__file__).parent.parent


def _iso():
    return datetime.now(timezone.utc).isoformat()


def _uid():
    return str(uuid.uuid4())


# ---------- Models ----------
class CommissionRule(BaseModel):
    id: str = Field(default_factory=_uid)
    brand: str = "Kazo"
    master_category: str
    sub_category: str
    gender: str = "Women"
    lower_limit: float
    upper_limit: float
    price_range: str
    commission_model: str = "Split Commission and Logistics"
    commission_pct: float
    active: bool = True


class FixedFeeSlab(BaseModel):
    id: str = Field(default_factory=_uid)
    aisp_lower: float
    aisp_upper: float
    label: str
    fixed_fee: float
    active: bool = True


class GTChargeCell(BaseModel):
    id: str = Field(default_factory=_uid)
    sub_category: str
    level: str
    price_range: str
    price_lower: float
    price_upper: float
    charge: float
    active: bool = True


class ReturnFeeCell(BaseModel):
    id: str = Field(default_factory=_uid)
    level: str
    zone: str
    fee: float
    active: bool = True


class SubCategoryLevel(BaseModel):
    id: str = Field(default_factory=_uid)
    sub_category: str
    level: str


class ToleranceConfig(BaseModel):
    absolute_inr: float = 1.0
    percentage: float = 0.5
    materiality_inr: float = 100.0


class TaxRates(BaseModel):
    gst_rate: float = 0.18
    tcs_rate: float = 0.005
    tds_rate: float = 0.001


class SettlementSettings(BaseModel):
    """Configurable defaults applied during calculation. NOT hardcoded — user editable."""
    default_zone_when_missing: Optional[str] = "Zonal"  # applied when zone == '-' or empty
    treat_dash_as_missing_zone: bool = True  # if True, '-' is treated as missing (uses default)
    apply_default_zone: bool = True  # master switch; if False, missing zone flags as unmapped


# ---------- Seed data (from KAZO Myntra commission file) ----------
FIXED_FEE_SEED = [
    {"aisp_lower": 0, "aisp_upper": 100, "label": "0-100", "fixed_fee": 27},
    {"aisp_lower": 101, "aisp_upper": 300, "label": "101-300", "fixed_fee": 27},
    {"aisp_lower": 301, "aisp_upper": 500, "label": "301-500", "fixed_fee": 27},
    {"aisp_lower": 501, "aisp_upper": 1000, "label": "501-1000", "fixed_fee": 27},
    {"aisp_lower": 1001, "aisp_upper": 2000, "label": "1001-2000", "fixed_fee": 45},
    {"aisp_lower": 2001, "aisp_upper": 10000000, "label": ">2000", "fixed_fee": 61},
]

RETURN_FEE_SEED = [
    ("Level 1", "Local", 91), ("Level 1", "Zonal", 112), ("Level 1", "National", 167),
    ("Level 2", "Local", 112), ("Level 2", "Zonal", 153), ("Level 2", "National", 218),
    ("Level 3", "Local", 142), ("Level 3", "Zonal", 194), ("Level 3", "National", 259),
    ("Level 4", "Local", 214), ("Level 4", "Zonal", 276), ("Level 4", "National", 331),
    ("Level 5", "Local", 460), ("Level 5", "Zonal", 542), ("Level 5", "National", 649),
]

# GT logistics rate matrix — per (level, price band). Sub-category maps to a level;
# the charge is looked up from level+band.
GT_LEVEL_SEED = {
    "Level 1": [(0, 100, "0-100", 0), (101, 300, "101-300", 59), (301, 500, "301-500", 59),
                (501, 1000, "501-1000", 94), (1001, 2000, "1001-2000", 171),
                (2001, 10000000, ">2000", 207)],
    "Level 2": [(0, 100, "0-100", 0), (101, 300, "101-300", 83), (301, 500, "301-500", 83),
                (501, 1000, "501-1000", 118), (1001, 2000, "1001-2000", 194),
                (2001, 10000000, ">2000", 230)],
    "Level 3": [(0, 100, "0-100", 0), (101, 300, "101-300", 100), (301, 500, "301-500", 106),
                (501, 1000, "501-1000", 148), (1001, 2000, "1001-2000", 230),
                (2001, 10000000, ">2000", 266)],
    "Level 4": [(0, 100, "0-100", 0), (101, 300, "101-300", 100), (301, 500, "301-500", 153),
                (501, 1000, "501-1000", 189), (1001, 2000, "1001-2000", 277),
                (2001, 10000000, ">2000", 313)],
    "Level 5": [(0, 100, "0-100", 0), (101, 300, "101-300", 150), (301, 500, "301-500", 200),
                (501, 1000, "501-1000", 240), (1001, 2000, "1001-2000", 330),
                (2001, 10000000, ">2000", 400)],
}

# Sub-category → Level mapping from KAZO GTA working sheet
SUBCAT_LEVEL_SEED = {
    "Wallets": "Level 1", "Tshirts": "Level 1", "T-Shirts": "Level 1",
    "Trousers": "Level 1", "Travel Accessory": "Level 1", "Tops": "Level 1",
    "Sweatshirts": "Level 2", "Sweaters": "Level 2", "Skirts": "Level 1",
    "Shrug": "Level 1", "Shorts": "Level 1", "Shirts": "Level 1",
    "Scarves": "Level 1", "Jumpsuit": "Level 1", "Jeggings": "Level 1",
    "Jeans": "Level 1", "Jackets": "Level 3", "Handbags": "Level 3",
    "Duffel Bag": "Level 2", "Dresses": "Level 1", "Coats": "Level 4",
    "Clutches": "Level 1", "Caps": "Level 1", "Blazers": "Level 4",
    "Belts": "Level 1", "Backpacks": "Level 3", "Track Pants": "Level 1",
    "Ring": "Level 1", "Necklace and Chains": "Level 1", "Headband": "Level 1",
    "Hair Accessory": "Level 2", "Brooch": "Level 1",
    "Accessory Gift Set": "Level 3", "Tunics": "Level 1",
    "Sunglasses": "Level 2", "Perfume and Body Mist": "Level 2",
    "Earrings": "Level 1", "Corset": "Level 1", "Clothing Set": "Level 3",
    "Bracelet": "Level 1", "Co-Ords": "Level 1", "Mufflers": "Level 1",
    "Stoles": "Level 1", "Laptop Bag": "Level 3", "Messenger Bag": "Level 3",
}


def _load_commission_rules_from_file():
    """Load authoritative commission rules from the JSON extracted from Myntra master file."""
    path = ROOT_DIR / "data_myntra_commission_rules.json"
    if not path.exists():
        raise RuntimeError(f"Commission rules file missing: {path}")
    with open(path) as f:
        rules = json.load(f)
    # Coverage helper: for every APPAREL sub-cat present in rules, ensure lowercase
    # variant is also matched at runtime — matching is case-insensitive already, so
    # we don't need duplicate rows. But make sure Jeggings is present since it's in sales.
    have = {(r["master_category"].upper(), r["sub_category"].lower()) for r in rules}
    if ("APPAREL", "jeggings") not in have:
        # Copy from "Jeans" pattern if present, else from a Level-1 default (0-500 @ 5%, >500 @ 8.5%)
        rules.append({"brand": "Kazo", "master_category": "APPAREL", "sub_category": "Jeggings",
                      "gender": "Women", "lower_limit": 0, "upper_limit": 300,
                      "price_range": "0-300", "commission_model": "Split Commission and Logistics", "commission_pct": 0.05})
        rules.append({"brand": "Kazo", "master_category": "APPAREL", "sub_category": "Jeggings",
                      "gender": "Women", "lower_limit": 300, "upper_limit": 500,
                      "price_range": "301-500", "commission_model": "Split Commission and Logistics", "commission_pct": 0.07})
        rules.append({"brand": "Kazo", "master_category": "APPAREL", "sub_category": "Jeggings",
                      "gender": "Women", "lower_limit": 500, "upper_limit": 10000000,
                      "price_range": ">500", "commission_model": "Split Commission and Logistics", "commission_pct": 0.085})
    return rules


async def seed_defaults(dbh):
    """Seed masters on first startup. Idempotent."""
    if await dbh.fixed_fees.count_documents({}) == 0:
        await dbh.fixed_fees.insert_many([{**x, "id": _uid(), "active": True} for x in FIXED_FEE_SEED])

    if await dbh.return_fees.count_documents({}) == 0:
        await dbh.return_fees.insert_many([
            {"id": _uid(), "level": lvl, "zone": z, "fee": f, "active": True}
            for lvl, z, f in RETURN_FEE_SEED
        ])

    if await dbh.gt_charges.count_documents({}) == 0:
        docs = []
        for sub, lvl in SUBCAT_LEVEL_SEED.items():
            for lo, hi, label, ch in GT_LEVEL_SEED[lvl]:
                docs.append({
                    "id": _uid(), "sub_category": sub, "level": lvl,
                    "price_range": label, "price_lower": lo, "price_upper": hi,
                    "charge": ch, "active": True,
                })
        await dbh.gt_charges.insert_many(docs)

    if await dbh.subcat_levels.count_documents({}) == 0:
        await dbh.subcat_levels.insert_many([
            {"id": _uid(), "sub_category": s, "level": l}
            for s, l in SUBCAT_LEVEL_SEED.items()
        ])

    if await dbh.commission_rules.count_documents({}) == 0:
        loaded = _load_commission_rules_from_file()
        await dbh.commission_rules.insert_many([
            {**r, "id": _uid(), "active": True} for r in loaded
        ])

    if await dbh.tolerances.count_documents({}) == 0:
        await dbh.tolerances.insert_one({
            "id": _uid(), "absolute_inr": 1.0, "percentage": 0.5, "materiality_inr": 100.0,
        })

    if await dbh.tax_rates.count_documents({}) == 0:
        await dbh.tax_rates.insert_one({
            "id": _uid(), "gst_rate": 0.18, "tcs_rate": 0.005, "tds_rate": 0.001,
        })

    if await dbh.settlement_settings.count_documents({}) == 0:
        await dbh.settlement_settings.insert_one({
            "id": _uid(),
            "default_zone_when_missing": "Zonal",
            "treat_dash_as_missing_zone": True,
            "apply_default_zone": True,
        })


# ---------- Endpoints ----------
def _clean(d):
    d = dict(d)
    d.pop("_id", None)
    return d


@router.get("/masters/commission-rules")
async def list_commission_rules():
    docs = await db.commission_rules.find({}).sort(
        [("master_category", 1), ("sub_category", 1), ("lower_limit", 1)]
    ).to_list(5000)
    return [_clean(d) for d in docs]


@router.post("/masters/commission-rules")
async def upsert_commission_rule(rule: CommissionRule, _user=Depends(require_admin)):
    d = rule.model_dump()
    await db.commission_rules.update_one({"id": d["id"]}, {"$set": d}, upsert=True)
    return _clean(d)


@router.delete("/masters/commission-rules/{rule_id}")
async def delete_commission_rule(rule_id: str, _user=Depends(require_admin)):
    await db.commission_rules.delete_one({"id": rule_id})
    return {"ok": True}


@router.get("/masters/fixed-fees")
async def list_fixed_fees():
    docs = await db.fixed_fees.find({}).sort("aisp_lower", 1).to_list(200)
    return [_clean(d) for d in docs]


@router.post("/masters/fixed-fees")
async def upsert_fixed_fee(fee: FixedFeeSlab, _user=Depends(require_admin)):
    d = fee.model_dump()
    await db.fixed_fees.update_one({"id": d["id"]}, {"$set": d}, upsert=True)
    return _clean(d)


@router.delete("/masters/fixed-fees/{fee_id}")
async def delete_fixed_fee(fee_id: str, _user=Depends(require_admin)):
    await db.fixed_fees.delete_one({"id": fee_id})
    return {"ok": True}


@router.get("/masters/gt-charges")
async def list_gt_charges():
    docs = await db.gt_charges.find({}).sort(
        [("sub_category", 1), ("price_lower", 1)]
    ).to_list(5000)
    return [_clean(d) for d in docs]


@router.post("/masters/gt-charges")
async def upsert_gt_charge(cell: GTChargeCell, _user=Depends(require_admin)):
    d = cell.model_dump()
    await db.gt_charges.update_one({"id": d["id"]}, {"$set": d}, upsert=True)
    return _clean(d)


@router.delete("/masters/gt-charges/{cell_id}")
async def delete_gt_charge(cell_id: str, _user=Depends(require_admin)):
    await db.gt_charges.delete_one({"id": cell_id})
    return {"ok": True}


@router.get("/masters/return-fees")
async def list_return_fees():
    docs = await db.return_fees.find({}).sort([("level", 1), ("zone", 1)]).to_list(200)
    return [_clean(d) for d in docs]


@router.post("/masters/return-fees")
async def upsert_return_fee(cell: ReturnFeeCell, _user=Depends(require_admin)):
    d = cell.model_dump()
    await db.return_fees.update_one({"id": d["id"]}, {"$set": d}, upsert=True)
    return _clean(d)


@router.get("/masters/subcat-levels")
async def list_subcat_levels():
    docs = await db.subcat_levels.find({}).sort("sub_category", 1).to_list(500)
    return [_clean(d) for d in docs]


@router.post("/masters/subcat-levels")
async def upsert_subcat_level(item: SubCategoryLevel, _user=Depends(require_admin)):
    d = item.model_dump()
    await db.subcat_levels.update_one({"id": d["id"]}, {"$set": d}, upsert=True)
    return _clean(d)


@router.get("/masters/tolerance")
async def get_tolerance():
    doc = await db.tolerances.find_one({}, {"_id": 0})
    if not doc:
        raise HTTPException(500, "Tolerance not configured. Seed did not run.")
    return doc


@router.post("/masters/tolerance")
async def set_tolerance(payload: ToleranceConfig, _user=Depends(require_admin)):
    d = payload.model_dump()
    await db.tolerances.update_one({}, {"$set": d}, upsert=True)
    return d


@router.get("/masters/tax-rates")
async def get_tax_rates():
    doc = await db.tax_rates.find_one({}, {"_id": 0})
    if not doc:
        raise HTTPException(500, "Tax rates not configured.")
    return doc


@router.post("/masters/tax-rates")
async def set_tax_rates(payload: TaxRates, _user=Depends(require_admin)):
    d = payload.model_dump()
    await db.tax_rates.update_one({}, {"$set": d}, upsert=True)
    return d


@router.get("/masters/settlement-settings")
async def get_settlement_settings():
    doc = await db.settlement_settings.find_one({}, {"_id": 0})
    if not doc:
        raise HTTPException(500, "Settlement settings not configured.")
    return doc


@router.post("/masters/settlement-settings")
async def set_settlement_settings(payload: SettlementSettings, _user=Depends(require_admin)):
    d = payload.model_dump()
    await db.settlement_settings.update_one({}, {"$set": d}, upsert=True)
    return d


@router.post("/masters/reset-defaults")
async def reset_defaults(_user=Depends(require_admin)):
    """Wipe all masters and re-seed from the source file. DESTRUCTIVE. Admin only."""
    await db.commission_rules.delete_many({})
    await db.fixed_fees.delete_many({})
    await db.gt_charges.delete_many({})
    await db.return_fees.delete_many({})
    await db.subcat_levels.delete_many({})
    await seed_defaults(db)
    from cache_utils import invalidate as _inv
    _inv()
    return {"ok": True, "message": "Masters reseeded from KAZO Myntra source file"}


# ---------- Configuration Export / Import ----------
_EXPORT_SPEC = [
    ("commission_rules", "Commission Rules",
     ["id", "brand", "master_category", "sub_category", "gender", "lower_limit", "upper_limit",
      "price_range", "commission_model", "commission_pct", "active"]),
    ("fixed_fees", "Fixed Fee",
     ["id", "aisp_lower", "aisp_upper", "label", "fixed_fee", "active"]),
    ("gt_charges", "GT Charges",
     ["id", "sub_category", "level", "price_range", "price_lower", "price_upper", "charge", "active"]),
    ("return_fees", "Return Fee",
     ["id", "level", "zone", "fee", "active"]),
    ("subcat_levels", "Sub-Cat Levels",
     ["id", "sub_category", "level"]),
    ("tolerances", "Tolerance",
     ["absolute_inr", "percentage", "materiality_inr"]),
    ("tax_rates", "Tax Rates",
     ["gst_rate", "tcs_rate", "tds_rate"]),
    ("settlement_settings", "Settlement Config",
     ["default_zone_when_missing", "treat_dash_as_missing_zone", "apply_default_zone"]),
]


@router.get("/masters/export")
async def export_masters():
    """Download all commission masters as a single multi-sheet Excel workbook."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for coll_name, sheet_name, cols in _EXPORT_SPEC:
        docs = await db[coll_name].find({}, {"_id": 0}).to_list(50000)
        ws = wb.create_sheet(sheet_name)
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.fill = header_fill
            cell.font = header_font
        for r, doc in enumerate(docs, start=2):
            for i, c in enumerate(cols, start=1):
                v = doc.get(c)
                if isinstance(v, (list, dict)):
                    v = str(v)
                ws.cell(row=r, column=i, value=v)
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, min(40, len(c) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="kazo-commission-masters-{now}.xlsx"'},
    )


@router.post("/masters/import")
async def import_masters(
    file: UploadFile = File(...),
    mode: str = "replace",  # replace | merge
    _user=Depends(require_admin),
):
    """Upload the Excel produced by /masters/export to bulk-update masters.

    mode='replace' wipes each collection before inserting (safer for full syncs).
    mode='merge'  upserts by id (or by natural key if id is blank).
    """
    from fastapi import UploadFile as _  # noqa: F401 (make sure symbol is bound)
    import io
    import openpyxl
    from cache_utils import invalidate as _inv

    if mode not in ("replace", "merge"):
        raise HTTPException(400, "mode must be 'replace' or 'merge'")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Not a valid Excel file: {e}")

    report: Dict[str, Any] = {"mode": mode, "sheets": []}
    for coll_name, sheet_name, cols in _EXPORT_SPEC:
        if sheet_name not in wb.sheetnames:
            report["sheets"].append({"sheet": sheet_name, "skipped": True, "reason": "not present"})
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            report["sheets"].append({"sheet": sheet_name, "skipped": True, "reason": "empty"})
            continue
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        col_idx = {c: header.index(c) for c in cols if c in header}
        if not col_idx:
            report["sheets"].append({"sheet": sheet_name, "skipped": True, "reason": "no matching columns"})
            continue

        docs: List[Dict[str, Any]] = []
        for r in rows[1:]:
            if not r or all(v is None for v in r):
                continue
            doc: Dict[str, Any] = {}
            for c, i in col_idx.items():
                v = r[i]
                if isinstance(v, str):
                    v = v.strip()
                doc[c] = v
            # Ensure id
            if not doc.get("id") and coll_name not in ("tolerances", "tax_rates", "settlement_settings"):
                doc["id"] = _uid()
            docs.append(doc)

        if coll_name in ("tolerances", "tax_rates", "settlement_settings"):
            # Singleton documents — take the first row and upsert
            if docs:
                await db[coll_name].delete_many({})
                await db[coll_name].insert_one(docs[0])
            report["sheets"].append({"sheet": sheet_name, "singleton": True, "count": 1 if docs else 0})
            continue

        if mode == "replace":
            await db[coll_name].delete_many({})
            if docs:
                await db[coll_name].insert_many(docs)
            report["sheets"].append({"sheet": sheet_name, "mode": "replace", "count": len(docs)})
        else:
            upserted = 0
            for d in docs:
                await db[coll_name].update_one({"id": d["id"]}, {"$set": d}, upsert=True)
                upserted += 1
            report["sheets"].append({"sheet": sheet_name, "mode": "merge", "count": upserted})

    _inv()
    return {"ok": True, **report}
