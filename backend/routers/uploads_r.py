"""File uploads: Sales data (Myntra Raw_Online Sale-m) and Settlement report."""
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from pydantic import BaseModel

from db import db
from cache_utils import invalidate as invalidate_cache


def _regex_escape(s: str) -> str:
    return re.escape(s)

router = APIRouter(tags=["uploads"])


def _iso():
    return datetime.now(timezone.utc).isoformat()


def _uid():
    return str(uuid.uuid4())


def _norm_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("₹", "").strip())
    except Exception:
        return 0.0


def _date_iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    try:
        return str(v)
    except Exception:
        return None


def _to_month(v) -> Optional[str]:
    """Convert 'Apr-26' / 'April 2026' / datetime / ISO string to 'YYYY-MM'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    for fmt in ("%b-%y", "%b-%Y", "%B-%y", "%B-%Y", "%B %Y", "%b %Y", "%Y-%m", "%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    # ISO date fallback: YYYY-MM-DDTHH:MM:SS...
    m = re.match(r"^(\d{4})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


# ---------- Sales upload ----------
# Canonical target fields + a list of possible source header variants.
# Match is case-insensitive and whitespace-insensitive.
# NOTE: Aliases cover Myntra, Amazon MTR/Settlement, AJIO, Nykaa, Tata Cliq,
# Flipkart. New portals just add their column labels here; the parser routes
# by `portal` query-param at ingestion time.
SALES_HEADER_ALIASES: Dict[str, List[str]] = {
    "order_date": ["Order Date", "OrderDate", "Order_Date", "purchase-date", "Purchase Date",
                    "Order Placed Date", "Invoice Date"],
    "txn_type": ["Txn Type", "Transaction Type", "Type", "transaction-type",
                  "Order Type"],
    "brand": ["Brand", "Brand Name"],
    "month": ["Month", "Report Month"],
    "posting_date": ["Posting Date", "PostingDate", "posted-date-time",
                      "Invoice Date"],
    "order_status": ["Order Status", "Status", "shipment-status", "order-status",
                      "Item Status"],
    "portal_name": ["Portal Name", "Portal", "Marketplace", "marketplace-name"],
    "sales_invoice_no": ["Sales Invoice No", "Invoice No", "Invoice Number", "invoice-number"],
    "online_order_id": ["Online Order Id", "Order Id", "Order ID", "OrderId",
                         "amazon-order-id", "Amazon Order ID", "amazon_order_id",
                         "Order Number", "Sub Order Number", "sub-order-id",
                         "Order Item Id", "amazon-order-item-id"],
    "sku": ["Sku", "SKU", "sku", "ASIN", "seller-sku", "item-sku",
             "Item Code", "Product SKU", "Style Id", "MSKU"],
    "zone": ["Shipped To _ZONE", "Zone", "Shipped Zone", "Shipping Zone",
              "ship-state", "ship-city", "State", "Buyer State",
              "Ship To State", "Delivery State"],
    "qty": ["QTY-Final", "Qty", "Quantity", "Order Qty", "quantity", "quantity-purchased", "quantity_purchased",
             "Item Qty", "Ordered Qty"],
    "mrp": ["MRP", "MRP/Item", "item-price", "Item Price", "Unit Price"],
    "total_mrp": ["Total MRP", "MRP Total"],
    "customer_discount": ["Cust. Discount", "Customer Discount", "Discount",
                           "item-promotion-discount", "promotion-discount"],
    "nsv_val": ["NSV VAL.", "NSV Value", "NSV Val", "NSV",
                 "Net Amount", "Net Payable", "Total Amount",
                 "product_sales", "Product Sales", "principal", "Principal Amount",
                 "Invoice Amount", "Order Item Total", "Sale Amount",
                 "Item Total Amount"],
    "nsv_per_unit": ["NSV/Unit", "NSV per Unit"],
    "main_category": ["Main Category", "Master Category", "Product Type", "product-type"],
    "category": ["Category", "Product Category", "item-category"],
    "sub_category": ["Sub Category_ GTA Charges", "Sub Category", "Sub-Category", "Subcategory",
                      "product-sub-category", "Product Sub-Category"],
    "posting_location_code": ["Posting Location Code", "Posting_Location Code", "Location Code",
                                "Warehouse Code", "Fulfilment Center", "ship-from-code"],
    "actual_gt_amount": ["GT Amount (Inc. gst)", "GT Amount", "GT Charges (Inc GST)",
                          "shipping-fee", "shipping_fee", "logistics-fee",
                          "Shipping Fee", "Logistics Fee"],
    "actual_fixed_fee": ["Fixed Fee-New", "Fixed Fee", "fba-fees", "fixed-fee",
                          "FBA Fee", "Fulfilment Fee", "Closing Fee"],
    "actual_return_fee": ["Return Fee-New", "Return Fee", "Return Shipping Fee"],
    "actual_commission_value": ["Commission Value.-New", "Commission Value", "Commission",
                                  "selling-fees", "referral-fee", "Referral Fee", "commission-fee",
                                  "Marketplace Fee"],
}


def _build_alias_lookup(aliases: Dict[str, List[str]]) -> Dict[str, str]:
    m = {}
    for canonical, variants in aliases.items():
        for v in variants:
            m[v.strip().lower()] = canonical
    return m


SALES_LOOKUP = _build_alias_lookup(SALES_HEADER_ALIASES)


# --------------------------------------------------------------------------
# Portal-specific post-parse normalisation.
# Non-Myntra portals use different vocabulary for txn_type / order_status.
# We normalise them into the canonical taxonomy used by _classify_order:
#   txn_type ∈ {Sales, Return}
#   order_status ∈ {Delivered, DTO, RTO, Internal Cancellation, ...}
# so the generic portal calc engine can classify each row correctly.
# --------------------------------------------------------------------------
def _normalize_row_for_portal(rec: Dict[str, Any], portal: str) -> Dict[str, Any]:
    """Map portal-native status vocabularies onto the canonical taxonomy.

    Mutates and returns `rec`. Safe no-op for Myntra (already canonical).
    """
    portal = (portal or "myntra").lower()
    if portal == "myntra":
        return rec

    raw_status = (rec.get("order_status") or "").strip().lower()
    raw_txn = (rec.get("txn_type") or "").strip().lower()
    nsv = rec.get("nsv_val") or 0

    # Amazon MTR / Settlement reports
    if portal == "amazon":
        # Refund rows come with negative principal or txn_type in {refund, return}
        if "refund" in raw_txn or "return" in raw_txn or nsv < 0:
            rec["txn_type"] = "Return"
            # cancellation on refund side → DTO / RTO. Amazon lumps them under Cancel.
            if "cancel" in raw_status:
                rec["order_status"] = "DTO"
            elif "rto" in raw_status or "returned" in raw_status:
                rec["order_status"] = "RTO"
            else:
                rec["order_status"] = "Delivered"
        else:
            rec["txn_type"] = "Sales"
            if "cancel" in raw_status:
                rec["order_status"] = "Internal Cancellation"
            elif raw_status.startswith("shipped") or raw_status.startswith("delivered"):
                rec["order_status"] = "Delivered"
        return rec

    # AJIO / Nykaa / Tata Cliq / Flipkart share a common convention:
    # 'Order' → Sales, 'Return' → Return, 'Cancel' → Internal Cancellation
    if "return" in raw_txn or "refund" in raw_txn or nsv < 0:
        rec["txn_type"] = "Return"
        if "cancel" in raw_status:
            rec["order_status"] = "DTO"
        elif "rto" in raw_status:
            rec["order_status"] = "RTO"
    else:
        rec["txn_type"] = "Sales"
        if "cancel" in raw_status:
            rec["order_status"] = "Internal Cancellation"
        elif not raw_status:
            rec["order_status"] = "Delivered"
    return rec


def _find_header_row(ws, lookup: Dict[str, str]):
    """Scan first 10 rows for a header row with the most matches against alias lookup."""
    best = (0, None, None)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else "" for v in row]
        matches = sum(1 for v in vals if v in lookup)
        if matches > best[0]:
            best = (matches, row_idx, row)
    if best[0] < 3:
        return None, None
    return best[1], best[2]


def _parse_sales_xlsx(content: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Rank sheets by header matches
    best_sheet = None
    best_score = 0
    best_header_row = None
    best_header = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        hr, header = _find_header_row(ws, SALES_LOOKUP)
        if not header:
            continue
        score = sum(1 for h in header if h and str(h).strip().lower() in SALES_LOOKUP)
        if score > best_score:
            best_sheet, best_score, best_header_row, best_header = sname, score, hr, header
    if not best_sheet:
        raise HTTPException(400, "Could not detect a valid sales sheet. Ensure headers include Order Id, SKU, MRP, NSV etc.")

    ws = wb[best_sheet]
    header_norm = [str(h).strip() if h else "" for h in best_header]
    col_idx: Dict[str, int] = {}
    for i, name in enumerate(header_norm):
        canonical = SALES_LOOKUP.get(name.lower())
        if canonical and canonical not in col_idx:
            col_idx[canonical] = i

    # Require the minimum set of columns. sub_category is optional for non-Myntra portals.
    required = {"online_order_id", "sku", "nsv_val"}
    missing = required - set(col_idx.keys())
    if missing:
        raise HTTPException(400, f"Missing required columns: {sorted(missing)}. Found: {sorted(col_idx.keys())}")

    def gv(row, k):
        return row[col_idx[k]] if k in col_idx else None

    accepted = []
    rejected = []
    for r_no, row in enumerate(ws.iter_rows(min_row=best_header_row + 1, values_only=True), start=best_header_row + 1):
        if not row or all(v is None for v in row):
            continue
        try:
            month_raw = gv(row, "month")
            posting = _date_iso(gv(row, "posting_date"))
            order_date_iso = _date_iso(gv(row, "order_date"))
            report_month = _to_month(month_raw) or _to_month(posting) or _to_month(order_date_iso)
            rec = {
                "order_date": order_date_iso,
                "txn_type": _norm_str(gv(row, "txn_type")),
                "brand": _norm_str(gv(row, "brand")),
                "month": _norm_str(month_raw),
                "report_month": report_month,
                "posting_date": posting,
                "order_status": _norm_str(gv(row, "order_status")),
                "portal_name": _norm_str(gv(row, "portal_name")),
                "sales_invoice_no": _norm_str(gv(row, "sales_invoice_no")),
                "online_order_id": _norm_str(gv(row, "online_order_id")),
                "sku": _norm_str(gv(row, "sku")),
                "zone": _norm_str(gv(row, "zone")),
                "qty": _num(gv(row, "qty")),
                "mrp": _num(gv(row, "mrp")),
                "total_mrp": _num(gv(row, "total_mrp")),
                "customer_discount": _num(gv(row, "customer_discount")),
                "nsv_val": _num(gv(row, "nsv_val")),
                "nsv_per_unit": _num(gv(row, "nsv_per_unit")),
                "main_category": _norm_str(gv(row, "main_category")),
                "category": _norm_str(gv(row, "category")),
                "sub_category": _norm_str(gv(row, "sub_category")),
                "posting_location_code": _norm_str(gv(row, "posting_location_code")),
                "actual_gt_amount": _num(gv(row, "actual_gt_amount")),
                "actual_fixed_fee": _num(gv(row, "actual_fixed_fee")),
                "actual_return_fee": _num(gv(row, "actual_return_fee")),
                "actual_commission_value": _num(gv(row, "actual_commission_value")),
                "_row_no": r_no,
            }
            if not rec.get("online_order_id") or not rec.get("sku") or rec["sku"] == "-":
                rejected.append({"row_no": r_no, "reason": "Missing/placeholder Online Order ID or SKU"})
                continue
            # Note: Returns come with negative NSV / QTY — accept them; the calculation
            # engine sign-normalizes based on txn_type/order_status.
            accepted.append(rec)
        except Exception as e:
            rejected.append({"row_no": r_no, "reason": f"Parse error: {e}"})

    return {"accepted": accepted, "rejected": rejected, "sheet": best_sheet, "header_row": best_header_row}


@router.post("/uploads/sales")
async def upload_sales(file: UploadFile = File(...), portal: str = Query("myntra")):
    portal = (portal or "myntra").strip().lower()
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx files are supported")
    content = await file.read()
    parsed = _parse_sales_xlsx(content)
    upload_id = _uid()
    total_accepted = len(parsed["accepted"])
    total_rejected = len(parsed["rejected"])

    months: Dict[str, int] = {}
    if parsed["accepted"]:
        docs = []
        for r in parsed["accepted"]:
            d = dict(r)
            d.pop("_row_no", None)
            d.update({
                "id": _uid(), "upload_id": upload_id,
                "uploaded_at": _iso(),
                "source_file": file.filename,
                "portal": portal,
            })
            # Normalise txn_type / order_status to the canonical vocabulary
            # so the calculation engine can classify each row deterministically.
            _normalize_row_for_portal(d, portal)
            if d.get("report_month"):
                months[d["report_month"]] = months.get(d["report_month"], 0) + 1
            docs.append(d)
        for i in range(0, len(docs), 1000):
            await db.sales.insert_many(docs[i:i + 1000])

    upload_doc = {
        "id": upload_id, "type": "sales", "filename": file.filename,
        "uploaded_at": _iso(), "sheet": parsed["sheet"],
        "accepted_count": total_accepted, "rejected_count": total_rejected,
        "rejections_sample": parsed["rejected"][:50],
        "months": months,
        "status": "processed",
        "portal": portal,
    }
    await db.uploads.insert_one({**upload_doc})
    # Store the raw file bytes so it can be re-downloaded (raw/original upload).
    # Base64 encode inside a separate collection to keep the uploads row small.
    import base64 as _b64
    await db.upload_files.insert_one({
        "upload_id": upload_id,
        "filename": file.filename,
        "content_b64": _b64.b64encode(content).decode("ascii"),
        "size": len(content),
        "portal": portal,
        "type": "sales",
        "uploaded_at": _iso(),
    })

    invalidate_cache()

    return {
        "upload_id": upload_id,
        "accepted_count": total_accepted,
        "rejected_count": total_rejected,
        "rejections_sample": parsed["rejected"][:20],
        "sheet": parsed["sheet"],
        "filename": file.filename,
        "months": months,
    }


# ---------- Settlement upload ----------
SETTLEMENT_HEADER_ALIASES: Dict[str, List[str]] = {
    "online_order_id": ["Order Id", "Order ID", "Online Order Id", "OrderId"],
    "sku": ["Sku", "SKU", "sku"],
    "settlement_date": ["Settlement Date", "Settled Date", "Payout Date"],
    "settled_commission": ["Commission", "Marketplace Fee", "Commission Amount", "Commission (incl GST)"],
    "settled_fixed_fee": ["Fixed Fee", "Closing Fee"],
    "settled_gt_charge": ["GT Charges", "Logistics", "Logistics Fee", "GT"],
    "settled_return_fee": ["Return Fee", "Return Shipping Fee"],
    "settled_tcs": ["TCS"],
    "settled_tds": ["TDS"],
    "settled_amount": ["Settlement Value", "Payout", "Net Settlement", "Settled Amount"],
    "selling_price": ["Selling Price", "SP"],
    "mrp": ["MRP"],
    "order_status": ["Order Status", "Status"],
    "zone": ["Zone", "Shipping Zone"],
    "month": ["Month", "Report Month"],
}
SETTLEMENT_LOOKUP = _build_alias_lookup(SETTLEMENT_HEADER_ALIASES)


def _parse_settlement_xlsx(content: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best_sheet = None
    best_score = 0
    best_header_row = None
    best_header = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        hr, header = _find_header_row(ws, SETTLEMENT_LOOKUP)
        if not header:
            continue
        score = sum(1 for h in header if h and str(h).strip().lower() in SETTLEMENT_LOOKUP)
        if score > best_score:
            best_sheet, best_score, best_header_row, best_header = sname, score, hr, header
    if not best_sheet:
        raise HTTPException(400, "Could not detect a valid settlement sheet. Ensure headers include Order Id, SKU, Commission, etc.")

    ws = wb[best_sheet]
    header_norm = [str(h).strip() if h else "" for h in best_header]
    col_idx: Dict[str, int] = {}
    for i, name in enumerate(header_norm):
        canonical = SETTLEMENT_LOOKUP.get(name.lower())
        if canonical and canonical not in col_idx:
            col_idx[canonical] = i

    required = {"online_order_id", "sku"}
    missing = required - set(col_idx.keys())
    if missing:
        raise HTTPException(400, f"Settlement file missing required columns: {sorted(missing)}")

    def gv(row, k):
        return row[col_idx[k]] if k in col_idx else None

    accepted = []
    rejected = []
    for r_no, row in enumerate(ws.iter_rows(min_row=best_header_row + 1, values_only=True), start=best_header_row + 1):
        if not row or all(v is None for v in row):
            continue
        try:
            settlement_date = _date_iso(gv(row, "settlement_date"))
            report_month = _to_month(gv(row, "month")) or _to_month(settlement_date)
            rec = {
                "online_order_id": _norm_str(gv(row, "online_order_id")),
                "sku": _norm_str(gv(row, "sku")),
                "settlement_date": settlement_date,
                "report_month": report_month,
                "settled_commission": _num(gv(row, "settled_commission")),
                "settled_fixed_fee": _num(gv(row, "settled_fixed_fee")),
                "settled_gt_charge": _num(gv(row, "settled_gt_charge")),
                "settled_return_fee": _num(gv(row, "settled_return_fee")),
                "settled_tcs": _num(gv(row, "settled_tcs")),
                "settled_tds": _num(gv(row, "settled_tds")),
                "settled_amount": _num(gv(row, "settled_amount")),
                "selling_price": _num(gv(row, "selling_price")),
                "zone": _norm_str(gv(row, "zone")),
                "order_status": _norm_str(gv(row, "order_status")),
                "_row_no": r_no,
            }
            if not rec["online_order_id"] or not rec["sku"]:
                rejected.append({"row_no": r_no, "reason": "Missing Order ID/SKU"})
                continue
            accepted.append(rec)
        except Exception as e:
            rejected.append({"row_no": r_no, "reason": f"Parse error: {e}"})
    return {"accepted": accepted, "rejected": rejected, "sheet": best_sheet, "header_row": best_header_row}


@router.post("/uploads/settlement")
async def upload_settlement(file: UploadFile = File(...), portal: str = Query("myntra")):
    portal = (portal or "myntra").strip().lower()
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx files are supported")
    content = await file.read()
    parsed = _parse_settlement_xlsx(content)
    upload_id = _uid()

    if parsed["accepted"]:
        docs = []
        for r in parsed["accepted"]:
            d = dict(r)
            d.pop("_row_no", None)
            d.update({
                "id": _uid(), "upload_id": upload_id,
                "uploaded_at": _iso(),
                "source_file": file.filename,
                "portal": portal,
            })
            docs.append(d)
        for i in range(0, len(docs), 1000):
            await db.settlement.insert_many(docs[i:i + 1000])

    upload_doc = {
        "id": upload_id, "type": "settlement", "filename": file.filename,
        "uploaded_at": _iso(), "sheet": parsed["sheet"],
        "accepted_count": len(parsed["accepted"]),
        "rejected_count": len(parsed["rejected"]),
        "rejections_sample": parsed["rejected"][:50],
        "status": "processed",
        "portal": portal,
    }
    await db.uploads.insert_one(upload_doc)
    # Store raw file bytes for re-download.
    import base64 as _b64
    await db.upload_files.insert_one({
        "upload_id": upload_id,
        "filename": file.filename,
        "content_b64": _b64.b64encode(content).decode("ascii"),
        "size": len(content),
        "portal": portal,
        "type": "settlement",
        "uploaded_at": _iso(),
    })

    invalidate_cache("periods")
    invalidate_cache("overview")

    return {
        "upload_id": upload_id,
        "accepted_count": len(parsed["accepted"]),
        "rejected_count": len(parsed["rejected"]),
        "rejections_sample": parsed["rejected"][:20],
        "sheet": parsed["sheet"],
        "filename": file.filename,
    }


@router.get("/uploads")
async def list_uploads(kind: Optional[str] = Query(None), portal: Optional[str] = Query(None)):
    q = {}
    if kind:
        q["type"] = kind
    if portal and portal.lower() != "all":
        q["portal"] = portal.lower()
    docs = await db.uploads.find(q, {"_id": 0}).sort("uploaded_at", -1).to_list(200)
    return docs


@router.delete("/uploads/{upload_id}")
async def delete_upload(upload_id: str):
    up = await db.uploads.find_one({"id": upload_id})
    if not up:
        raise HTTPException(404, "Upload not found")
    coll = "sales" if up["type"] == "sales" else "settlement"
    await db[coll].delete_many({"upload_id": upload_id})
    if coll == "sales":
        # also delete related calculations
        sales_ids = [s["id"] async for s in db.sales.find({"upload_id": upload_id}, {"id": 1})]
        if sales_ids:
            await db.calculations.delete_many({"sales_id": {"$in": sales_ids}})
    await db.uploads.delete_one({"id": upload_id})
    await db.upload_files.delete_one({"upload_id": upload_id})
    invalidate_cache()  # nuke all — sales/calc/disc caches all potentially stale
    return {"ok": True}


@router.get("/uploads/{upload_id}/download")
async def download_upload(upload_id: str):
    """Return the original XLSX bytes that were uploaded. Raw file is stored
    at upload time in db.upload_files. Older uploads (pre-2026-02) predate
    the raw-file store and will 404 with a helpful message."""
    doc = await db.upload_files.find_one({"upload_id": upload_id})
    if not doc:
        up = await db.uploads.find_one({"id": upload_id})
        if not up:
            raise HTTPException(404, "Upload not found")
        raise HTTPException(
            410,
            "Raw file for this upload wasn't captured (uploaded before raw-file store was enabled). Re-upload the file to enable download.",
        )
    import base64 as _b64
    from fastapi.responses import Response
    raw = _b64.b64decode(doc["content_b64"])
    return Response(
        content=raw,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{doc["filename"]}"',
            "Content-Length": str(len(raw)),
        },
    )


@router.get("/sales")
async def list_sales(
    upload_id: Optional[str] = None,
    report_month: Optional[str] = None,
    period_type: Optional[str] = None,
    period_value: Optional[str] = None,
    portal: Optional[str] = None,
    search: Optional[str] = None,
    zone: Optional[str] = None,
    category: Optional[str] = None,
    sub_category: Optional[str] = None,
    main_category: Optional[str] = None,
    order_status: Optional[str] = None,
    txn_type: Optional[str] = None,
    limit: int = Query(200, le=2000),
    skip: int = 0,
    sort_by: str = "order_date",
    sort_dir: str = "desc",
):
    from period_utils import month_query as _mq
    q: Dict[str, Any] = {}
    if period_type:
        q.update(_mq(period_type, period_value))
    elif report_month:
        q["report_month"] = report_month
    if portal and portal.lower() != "all":
        q["portal"] = portal.lower()
    if upload_id:
        q["upload_id"] = upload_id
    if zone:
        q["zone"] = zone
    if category:
        q["category"] = category
    if sub_category:
        q["sub_category"] = sub_category
    if main_category:
        q["main_category"] = main_category
    if order_status:
        q["order_status"] = order_status
    if txn_type:
        q["txn_type"] = txn_type
    if search:
        s = _regex_escape(search.strip())
        q["$or"] = [
            {"online_order_id": {"$regex": s, "$options": "i"}},
            {"sku": {"$regex": s, "$options": "i"}},
            {"sales_invoice_no": {"$regex": s, "$options": "i"}},
        ]
    total = await db.sales.count_documents(q)
    sort_map = {
        "order_date": "order_date", "sku": "sku", "order_id": "online_order_id",
        "nsv": "nsv_val", "mrp": "mrp", "qty": "qty", "month": "report_month",
        "sub_category": "sub_category", "zone": "zone",
    }
    sort_field = sort_map.get(sort_by, "order_date")
    direction = -1 if sort_dir == "desc" else 1
    docs = await db.sales.find(q, {"_id": 0}).sort(sort_field, direction).skip(skip).limit(limit).to_list(limit)
    return {"total": total, "items": docs}


@router.get("/sales/months")
async def list_sales_months():
    """Distinct report_month values present in the sales collection."""
    pipeline = [
        {"$match": {"report_month": {"$ne": None}}},
        {"$group": {"_id": "$report_month", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]
    rows = await db.sales.aggregate(pipeline).to_list(200)
    return [{"month": r["_id"], "count": r["count"]} for r in rows]


@router.get("/sales/summary")
async def sales_summary(
    upload_id: Optional[str] = None,
    report_month: Optional[str] = None,
    period_type: Optional[str] = None,
    period_value: Optional[str] = None,
    portal: Optional[str] = None,
    search: Optional[str] = None,
    zone: Optional[str] = None,
    category: Optional[str] = None,
    sub_category: Optional[str] = None,
    main_category: Optional[str] = None,
    order_status: Optional[str] = None,
    txn_type: Optional[str] = None,
):
    """Aggregate KPIs for the Sales Ledger. Order Qty = net qty (Sales − Returns)."""
    from period_utils import month_query as _mq
    q: Dict[str, Any] = {}
    if period_type:
        q.update(_mq(period_type, period_value))
    elif report_month:
        q["report_month"] = report_month
    if portal and portal.lower() != "all":
        q["portal"] = portal.lower()
    if upload_id:
        q["upload_id"] = upload_id
    if zone:
        q["zone"] = zone
    if category:
        q["category"] = category
    if sub_category:
        q["sub_category"] = sub_category
    if main_category:
        q["main_category"] = main_category
    if order_status:
        q["order_status"] = order_status
    if txn_type:
        q["txn_type"] = txn_type
    if search:
        s = _regex_escape(search.strip())
        q["$or"] = [
            {"online_order_id": {"$regex": s, "$options": "i"}},
            {"sku": {"$regex": s, "$options": "i"}},
            {"sales_invoice_no": {"$regex": s, "$options": "i"}},
        ]
    pipeline = [
        {"$match": q},
        {"$group": {
            "_id": None,
            "row_count": {"$sum": 1},
            "sales_rows": {"$sum": {"$cond": [{"$in": ["$txn_type", ["Sales", "sales", None]]}, 1, 0]}},
            "return_rows": {"$sum": {"$cond": [{"$in": ["$txn_type", ["Return", "return"]]}, 1, 0]}},
            "net_qty": {"$sum": {"$ifNull": ["$qty", 0]}},
            "nsv_total": {"$sum": {"$ifNull": ["$nsv_val", 0]}},
        }},
    ]
    rows = await db.sales.aggregate(pipeline).to_list(1)
    if not rows:
        return {"row_count": 0, "sales_rows": 0, "return_rows": 0, "net_qty": 0, "net_orders": 0, "nsv_total": 0}
    r = rows[0]
    return {
        "row_count": r.get("row_count", 0),
        "sales_rows": r.get("sales_rows", 0),
        "return_rows": r.get("return_rows", 0),
        "net_orders": r.get("sales_rows", 0) - r.get("return_rows", 0),
        "net_qty": round(r.get("net_qty", 0) or 0, 2),
        "nsv_total": round(r.get("nsv_total", 0) or 0, 2),
    }


@router.get("/sales/export")
async def export_sales_xlsx(
    upload_id: Optional[str] = None,
    report_month: Optional[str] = None,
    period_type: Optional[str] = None,
    period_value: Optional[str] = None,
    portal: Optional[str] = None,
    search: Optional[str] = None,
    zone: Optional[str] = None,
    sub_category: Optional[str] = None,
    main_category: Optional[str] = None,
    order_status: Optional[str] = None,
    txn_type: Optional[str] = None,
):
    """Export the Sales Ledger as an Excel workbook with all client-requested
    columns, including Brand, Sale Type, Posting Date, Item No, Posting_Location
    Code, Main Ctg, Level No, Price Range (NSV), Price Range (NSV after GT) — the
    last three are joined from the calculations collection.
    """
    from period_utils import month_query as _mq
    from fastapi.responses import StreamingResponse
    q: Dict[str, Any] = {}
    if period_type:
        q.update(_mq(period_type, period_value))
    elif report_month:
        q["report_month"] = report_month
    if portal and portal.lower() != "all":
        q["portal"] = portal.lower()
    if upload_id:
        q["upload_id"] = upload_id
    if zone:
        q["zone"] = zone
    if sub_category:
        q["sub_category"] = sub_category
    if main_category:
        q["main_category"] = main_category
    if order_status:
        q["order_status"] = order_status
    if txn_type:
        q["txn_type"] = txn_type
    if search:
        s = _regex_escape(search.strip())
        q["$or"] = [
            {"online_order_id": {"$regex": s, "$options": "i"}},
            {"sku": {"$regex": s, "$options": "i"}},
            {"sales_invoice_no": {"$regex": s, "$options": "i"}},
        ]

    # Stream sales rows (up to 200k for safety)
    sales_cursor = db.sales.find(q, {"_id": 0}).limit(200000)
    sale_docs = [s async for s in sales_cursor]
    sales_ids = [s["id"] for s in sale_docs]

    # Pull matching calc rows for the joined columns (Level, Price ranges, Commission %)
    calc_map: Dict[str, Dict[str, Any]] = {}
    if sales_ids:
        async for c in db.calculations.find({"sales_id": {"$in": sales_ids}}, {"_id": 0}):
            calc_map[c["sales_id"]] = c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Ledger"
    headers = [
        "Portal", "Brand", "Sale Type", "Order Status", "Posting Date", "Report Month",
        "Online Order ID", "Item No (SKU)", "Sales Invoice No",
        "Posting_Location Code", "Main Ctg", "Category", "Sub Category", "Level No",
        "Zone", "Qty", "MRP", "Total MRP", "Customer Discount",
        "NSV Value", "NSV per Unit", "NSV after GT",
        "Price Range - Key (NSV)", "Price Range - Key (NSV after GT)",
        "Commission %", "Commission (Expected)", "GT Charge (Expected)",
        "Fixed Fee (Expected)", "Return Fee (Expected)", "Total Deductions",
        "Expected Settlement",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    def _n(v):
        try:
            return round(float(v), 2) if v is not None and v != "" else None
        except Exception:
            return None

    for s in sale_docs:
        c = calc_map.get(s["id"], {}) or {}
        bd = c.get("breakdown") or {}
        crule = bd.get("commission_rule") or {}
        gt_cell = bd.get("gt_charge_cell") or {}
        row = [
            (s.get("portal") or "").upper(),
            s.get("brand"),
            s.get("txn_type"),
            s.get("order_status"),
            s.get("posting_date"),
            s.get("report_month"),
            s.get("online_order_id"),
            s.get("sku"),
            s.get("sales_invoice_no"),
            s.get("posting_location_code"),
            s.get("main_category"),
            s.get("category"),
            s.get("sub_category"),
            bd.get("level"),
            bd.get("zone") or s.get("zone"),
            _n(s.get("qty")),
            _n(s.get("mrp")),
            _n(s.get("total_mrp")),
            _n(s.get("customer_discount")),
            _n(s.get("nsv_val")),
            _n(s.get("nsv_per_unit")),
            _n(c.get("nsv_after_gt")),
            crule.get("price_range"),
            gt_cell.get("price_range"),
            crule.get("commission_pct"),
            _n(c.get("commission_incl_gst")),
            _n(c.get("gt_charge")),
            _n(c.get("fixed_fee_incl_gst")),
            _n(c.get("return_fee")),
            _n(c.get("total_deductions")),
            _n(c.get("expected_settlement")),
        ]
        ws.append(row)

    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fundle-sales-ledger-{report_month or period_value or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/settlement")
async def list_settlement(
    upload_id: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = Query(200, le=1000),
    skip: int = 0,
):
    q: Dict[str, Any] = {}
    if upload_id:
        q["upload_id"] = upload_id
    if search:
        q["$or"] = [
            {"online_order_id": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
        ]
    total = await db.settlement.count_documents(q)
    docs = await db.settlement.find(q, {"_id": 0}).sort("uploaded_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"total": total, "items": docs}
